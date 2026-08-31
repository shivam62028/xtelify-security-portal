__author__ = "richyrik"

import os, json, re, time, zipfile, io
import pandas as pd
import httpx
from io import BytesIO
from datetime import datetime, timedelta, timezone, date
from fastapi import FastAPI, Request, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, ORJSONResponse
import smtplib
from email.message import EmailMessage
import base64
from pydantic import BaseModel
from typing import Optional
from openpyxl import load_workbook
from functools import lru_cache
import hashlib
from pymongo import MongoClient, UpdateOne

client = MongoClient(
    "mongodb://127.0.0.1:27017/",
    serverSelectionTimeoutMS=5000
)
mongo_db = client["xtelify_db"]
issues_collection = mongo_db["vulnerabilities"]
upload_history_collection = mongo_db["upload_history"]
ai_remediation_cache_collection = mongo_db["ai_remediation_cache"]

# In-memory cache for faster loading
_db_cache = None
_db_cache_time = 0
_db_cache_hash = None
CACHE_TTL = 5  # seconds

OLLAMA_URL = "http://127.0.0.1:11434/api/generate"
OLLAMA_MODEL = "llama3"

EXPECTED_COLUMNS = {
    "id", "name", "severity", "findingstatus", "score", "wizurl",
    "vendorseverity", "cvssseverity", "hasexploit", "hascisakeknownexploit",
    "firstdetected", "lastdetected", "resolvedat", "resolution", "remediation",
    "locationpath", "detailedname", "version", "fixedversion", "status",
    "subscriptionid", "subscriptionname", "namespaces", "clusters", "imageid",
    "cloudprovider", "cloud_provider", "accountid", "account_id", "accountname", "account_name",
    "resourcetype", "resource_type", "findingtypeid", "finding_type_id", "findingname", "finding_name",
    "resourceid", "resource_id", "resourcename", "resource_name", "compliancetags", "compliance_tags",
    "riskscore", "risk_score", "impact", "remediationtype", "remediation_type", "region",
    "issuekey", "issue key", "summary", "applicationname", "application name",
    "criticalitystatus", "criticality status", "reportedon", "reported on", "ageing",
    "assignee", "multipleassignee", "multiple assignee", "applicationowner", "application owner",
    "expectedtimeline", "expected timeline", "compliant", "non-compliant"
}

HIGH_SCORE_COLS = {"id", "name", "severity", "findingstatus", "issuekey", "issue key", "cloud_provider", "finding_name"}
MED_SCORE_COLS = {"score", "cvssseverity", "wizurl", "account_name", "resource_name", "applicationname", "summary"}
NEGATIVE_PATTERNS = ["grand total", "count of", "pivot", "impacted resources", "summary", "row labels", "column labels", "values", "total"]

PIVOT_INDICATORS = ["count of", "sum of", "average of", "row labels", "column labels", "grand total", "values"]

ALLOWED_LOB = ["wynk"]

CONTAINER_COLUMNS = {
    "wizurl", "cvssseverity", "hasexploit", "hascisaknownexploit", "findingstatus",
    "vendorseverity", "nvdseverity", "firstdetected", "lastdetected", "resolvedat",
    "detailedname", "fixedversion", "detectionlink", "assetregion", "provideruniqueid",
    "cloudprovider", "cloudplatform", "subscriptionid", "subscriptionname", "subscriptiontags",
    "executioncontext", "namespaces", "clusters", "imageid", "locationpath"
}

CSPM_COLUMNS = {
    "cloud_provider", "account_id", "account_name", "resource_type", "finding_type_id",
    "finding_name", "resource_id", "resource_name", "compliance_tags", "risk_score",
    "remediation_type", "region"
}

SAST_DAST_COLUMNS = {
    "issue key", "issuekey", "application name", "criticality status", "reported on",
    "ageing", "compliant/non-compliant", "expected timeline", "assignee",
    "multiple assignee", "application owner"
}

VAPT_COLUMNS = {
    "hostname", "ip", "protocol", "port", "risk factor", "uuid", "vprscore", "priority",
    "cve number", "vulnerability name", "vulnerability description", "solution",
    "vulnerability path", "vulnerability id", "vulnerability family", "repo name",
    "quarter", "lob name", "application name", "application owner", "vulnerability status",
    "vulnpubdate", "patchpubdate", "pluginpubdate", "pluginmoddate", "firstseen", "lastseen",
    "vulnerability type", "internet/non-intern"
}

def detect_file_format(columns):
    cols_lower = {str(c).lower().strip() for c in columns}
    cols_normalized = {str(c).lower().replace(" ", "").replace("_", "").replace("-", "").strip() for c in columns}
    cols_str = ' '.join(cols_lower)

    vapt_matches = 0
    for col in cols_lower:
        if col == "uuid":
            vapt_matches += 15
        if "vulnerability name" in col:
            vapt_matches += 10
        if "vulnerability description" in col:
            vapt_matches += 10
        if "vulnerability path" in col:
            vapt_matches += 8
        if "vulnerability family" in col:
            vapt_matches += 8
        if "vulnerability id" in col:
            vapt_matches += 8
        if "vulnerability status" in col:
            vapt_matches += 8
        if col == "vprscore" or "vpr" in col:
            vapt_matches += 8
        if "risk factor" in col:
            vapt_matches += 5
        if col == "hostname":
            vapt_matches += 5
        if "cve number" in col:
            vapt_matches += 5
        if col == "lastseen" or "last seen" in col:
            vapt_matches += 3
        if col == "firstseen" or "first seen" in col:
            vapt_matches += 3

    sast_dast_matches = 0
    for col in cols_lower:
        if "issue key" in col or col == "issuekey" or col == "issue_key":
            sast_dast_matches += 15
        if "criticality" in col or "criticality status" in col:
            sast_dast_matches += 8
        if "ageing" in col or "aging" in col:
            sast_dast_matches += 8
        if "expected timeline" in col or "timeline" in col:
            sast_dast_matches += 5
        if "compliant" in col or "non-compliant" in col:
            sast_dast_matches += 5
        if col == "assignee" or "multiple assignee" in col:
            sast_dast_matches += 5
        if col == "summary" and "issue key" in cols_str:
            sast_dast_matches += 5
        if "reported on" in col or "reported" in col:
            sast_dast_matches += 3

    cspm_matches = 0
    for col in cols_lower:
        if col == "cloud_provider" or col == "cloudprovider":
            cspm_matches += 10
        if col == "account_id" or col == "accountid":
            cspm_matches += 5
        if col == "account_name" or col == "accountname":
            cspm_matches += 5
        if col == "finding_type_id" or col == "findingtypeid":
            cspm_matches += 5
        if col == "finding_name" or col == "findingname":
            cspm_matches += 5
        if col == "resource_type" or col == "resourcetype":
            cspm_matches += 3
        if col == "compliance_tags" or col == "compliancetags":
            cspm_matches += 3
        if col == "risk_score" or col == "riskscore":
            cspm_matches += 3

    container_matches = 0
    for col in cols_lower:
        if col == "wizurl" or "wizurl" in col:
            container_matches += 10
        if col == "cvssseverity":
            container_matches += 5
        if col == "hasexploit":
            container_matches += 5
        if col == "findingstatus":
            container_matches += 5
        if col == "subscriptionid":
            container_matches += 5
        if col == "subscriptionname":
            container_matches += 5
        if col == "detailedname":
            container_matches += 3
        if col == "fixedversion":
            container_matches += 3
        if col == "namespaces" or col == "clusters":
            container_matches += 3
        if col == "imageid":
            container_matches += 5

    print(f"Format detection - Container: {container_matches}, CSPM: {cspm_matches}, SAST_DAST: {sast_dast_matches}, VAPT: {vapt_matches}")

    if vapt_matches > sast_dast_matches and vapt_matches > cspm_matches and vapt_matches > container_matches:
        return "VAPT"
    elif sast_dast_matches > cspm_matches and sast_dast_matches > container_matches:
        return "SAST_DAST"
    elif cspm_matches > container_matches:
        return "CSPM"
    else:
        return "CONTAINER"

POD_OWNER_MAPPING = {
    "xstream": "Shreya",
    "xstrm": "Shreya",
    "x-stream": "Shreya",
    "x_stream": "Shreya",
    "adtech": "Satya",
    "ad-tech": "Satya",
    "ad_tech": "Satya",
    "music": "Aakash",
    "wcf": "Yash",
    "vmax": "Dheeraj",
    "v-max": "Dheeraj",
    "v_max": "Dheeraj",
    "iptv-be": "Shreya",
    "iptv_be": "Shreya",
    "iptvbe": "Shreya",
    "iptv-backend": "Shreya",
    "data platform": "Abhinav/Vinod",
    "dataplatform": "Abhinav/Vinod",
    "data_platform": "Abhinav/Vinod",
    "data-platform": "Abhinav/Vinod",
    "msp": "Yash",
    "search": "Mohit",
    "ml": "Nisha",
    "catalog": "Aakash",
    "catalogue": "Aakash",
    "channels": "Vinod",
    "channel": "Vinod",
    "uclm": "Dheeraj/Satya",
    "iptv": "Anshu",
    "ktv": "Anshu",
    "discovery": "Aakash",
    "disco": "Aakash",
    "dp": "Vinod",
    "cmn": "Shiv Kumar",
    "ds": "Shiv Kumar",
}

CSPM_POD_KEYWORDS = [
    ("xstream", "Shreya"),
    ("xstrm", "Shreya"),
    ("x-stream", "Shreya"),
    ("x_stream", "Shreya"),
    ("adtech", "Satya"),
    ("music", "Aakash"),
    ("wcf", "Yash"),
    ("vmax", "Dheeraj"),
    ("iptv-be", "Shreya"),
    ("iptvbe", "Shreya"),
    ("data-platform", "Abhinav/Vinod"),
    ("dataplatform", "Abhinav/Vinod"),
    ("msp", "Yash"),
    ("search", "Mohit"),
    ("ml", "Nisha"),
    ("catalog", "Aakash"),
    ("channels", "Vinod"),
    ("channel", "Vinod"),
    ("uclm", "Dheeraj/Satya"),
    ("iptv", "Anshu"),
    ("ktv", "Anshu"),
    ("discovery", "Aakash"),
    ("disco", "Aakash"),
    ("cmn", "Shiv Kumar"),
    ("ds", "Shiv Kumar"),
    ("dp", "Vinod"),
    ("infra", "Shiv Kumar"),
    ("monitoring", "Shiv Kumar"),
    ("focus", "Shiv Kumar"),
    ("billingexp", "Shiv Kumar"),
    ("intc", "Anshu"),
    ("pre-music", "Aakash"),
    ("pre-vmax", "Dheeraj"),
    ("stg-msp", "Yash"),
    ("prd-channel", "Vinod"),
]

def get_cspm_pod_owner(*args):
    """
    Auto-assign CSPM issues based on account_id and account_name matching POD keywords.
    Returns the POD Owner name or 'Unassigned' if no match.
    """
    combined = " ".join([str(arg).lower() for arg in args if arg and str(arg).lower() not in ["", "na", "none", "nan"]])

    for keyword, owner in CSPM_POD_KEYWORDS:
        if keyword.lower() in combined:
            return owner

    return "Unassigned"

def generate_short_description(vuln_name, cve_id, severity, asset_type, detailed_name):
    desc_parts = []
    severity_words = {
        "critical": "Critical security flaw",
        "high": "High-risk vulnerability",
        "medium": "Moderate security issue",
        "low": "Minor security concern",
        "info": "Informational finding"
    }
    sev_lower = (severity or "medium").lower()
    sev_prefix = severity_words.get(sev_lower, "Security issue")
    name_lower = (vuln_name or "").lower()
    detailed_lower = (detailed_name or "").lower()
    combined = name_lower + " " + detailed_lower

    vuln_type = ""
    if any(x in combined for x in ["rce", "remote code", "command injection", "code execution"]):
        vuln_type = "allows remote code execution"
    elif any(x in combined for x in ["sql injection", "sqli", "sql inj"]):
        vuln_type = "SQL injection vulnerability found"
    elif any(x in combined for x in ["xss", "cross-site script", "cross site script"]):
        vuln_type = "cross-site scripting detected"
    elif any(x in combined for x in ["buffer overflow", "overflow", "memory corrupt"]):
        vuln_type = "memory corruption vulnerability"
    elif any(x in combined for x in ["dos", "denial of service", "denial-of-service"]):
        vuln_type = "denial of service possible"
    elif any(x in combined for x in ["auth", "authentication", "bypass", "privilege"]):
        vuln_type = "authentication bypass risk"
    elif any(x in combined for x in ["path traversal", "directory traversal", "lfi", "rfi"]):
        vuln_type = "path traversal vulnerability"
    elif any(x in combined for x in ["ssrf", "server-side request"]):
        vuln_type = "server-side request forgery"
    elif any(x in combined for x in ["xxe", "xml external"]):
        vuln_type = "XML external entity attack"
    elif any(x in combined for x in ["deserializ", "unserializ"]):
        vuln_type = "insecure deserialization flaw"
    elif any(x in combined for x in ["crypto", "encrypt", "ssl", "tls", "certificate"]):
        vuln_type = "cryptographic weakness detected"
    elif any(x in combined for x in ["config", "misconfig", "default", "hardcoded"]):
        vuln_type = "configuration issue found"
    elif any(x in combined for x in ["outdated", "upgrade", "version", "update", "patch"]):
        vuln_type = "outdated component needs update"
    elif any(x in combined for x in ["exposure", "leak", "sensitive", "disclosure"]):
        vuln_type = "information disclosure risk"
    elif any(x in combined for x in ["inject", "input valid"]):
        vuln_type = "injection vulnerability detected"
    elif any(x in combined for x in ["container", "docker", "kubernetes", "k8s", "image"]):
        vuln_type = "container security issue"
    elif any(x in combined for x in ["permission", "access control", "rbac"]):
        vuln_type = "access control weakness"
    elif any(x in combined for x in ["log4j", "log4shell"]):
        vuln_type = "Log4j vulnerability detected"
    elif any(x in combined for x in ["spring", "spring4shell"]):
        vuln_type = "Spring framework vulnerability"

    if vuln_type:
        # Use detected type
        if cve_id and cve_id.upper().startswith("CVE"):
            desc = f"{sev_prefix}: {vuln_type}"
        else:
            desc = f"{sev_prefix} - {vuln_type}"
    elif cve_id and cve_id.upper().startswith("CVE"):
        # Use CVE if no type detected
        desc = f"{sev_prefix} in {cve_id}"
    elif vuln_name:
        # Use first few words of name
        words = vuln_name.split()[:4]
        short_name = " ".join(words)
        desc = f"{sev_prefix}: {short_name}"
    else:
        # Fallback
        asset = asset_type if asset_type and asset_type not in ["", "NA"] else "system"
        desc = f"{sev_prefix} affecting {asset}"

    # Ensure 5-7 words (trim if too long)
    final_words = desc.split()
    if len(final_words) > 8:
        desc = " ".join(final_words[:7]) + "..."

    return desc


def get_pod_owner(*args):
    """
    Auto-detect POD owner from subscription name or ID.
    Matches keywords from POD_OWNER_MAPPING with smart matching.
    """
    search_text = " ".join([str(arg).lower() for arg in args if arg and str(arg).lower() not in ["na", "none", "nan", ""] ])

    if not search_text.strip():
        return ""

    normalized = search_text.replace("-", " ").replace("_", " ").replace(".", " ")
    sorted_keywords = sorted(POD_OWNER_MAPPING.keys(), key=len, reverse=True)

    for pod_keyword in sorted_keywords:
        owner = POD_OWNER_MAPPING[pod_keyword]
        if pod_keyword in search_text or pod_keyword in normalized or pod_keyword in normalized.split():
            return owner

    return ""


def ask_ollama_is_valid_row(row_sample):
    """Ask Ollama if a row is a valid vulnerability record (for edge cases)"""
    import requests
    try:
        prompt = f"""Is this a valid vulnerability/security finding record or a pivot table/summary row?

Row data: {row_sample}

Answer only: VALID or SKIP"""

        response = requests.post(
            OLLAMA_URL,
            json={"model": OLLAMA_MODEL, "prompt": prompt, "stream": False},
            timeout=5
        )
        if response.status_code == 200:
            answer = response.json().get("response", "").strip().upper()
            return "VALID" in answer
    except:
        pass
    return True  # Default to valid if Ollama fails


def is_pivot_or_summary_row(row, use_ollama_for_edge_cases=False):
    """Skip pivot table rows, summary rows, blank rows, and count rows

    Hybrid approach:
    1. Quick keyword filter (catches 95%)
    2. If uncertain, ask Ollama to verify (disabled by default for performance)
    """
    row_values = [str(v).strip().lower() for v in row.values() if v is not None]
    row_str = ' '.join(row_values)

    # DEFINITE SKIP - keyword match (fast)
    skip_keywords = ['grand total', 'row labels', 'count of', 'sum of', 'average of',
                   '(blank)', 'total', 'subtotal', 'pivot', 'summary']
    if any(kw in row_str for kw in skip_keywords):
        return True

    # DEFINITE SKIP - too few values
    non_empty = [v for v in row_values if v and v not in ['nan', 'none', 'na', '']]
    if len(non_empty) < 3:
        return True

    # DEFINITE SKIP - blank/invalid ID
    # Check various ID column names (case-insensitive)
    id_col = None
    for key in row.keys():
        key_lower = str(key).lower().replace(" ", "").replace("_", "")
        if key_lower in ['id', 'issueid', 'issuekey', 'issue_key']:
            id_col = row.get(key)
            break

    if id_col:
        id_str = str(id_col).strip().lower()
        if id_str in ['', 'nan', 'none', 'na', 'null'] or id_str.startswith('('):
            return True

    # DEFINITE VALID - has CVE pattern or looks like a valid issue key
    if 'cve-' in row_str:
        return False

    # Check if any value looks like a SAST_DAST issue key (e.g., contains letters and numbers)
    for v in row_values:
        if v and len(v) > 3 and any(c.isalpha() for c in v) and any(c.isdigit() for c in v):
            if not any(skip in v for skip in ['nan', 'none', 'total', 'count']):
                return False  # Looks like valid data

    # EDGE CASE - uncertain, ask Ollama
    uncertain_patterns = ['count', 'total', 'blank', 'label', 'header']
    is_uncertain = any(p in row_str for p in uncertain_patterns)

    if is_uncertain and use_ollama_for_edge_cases:
        row_sample = dict(list(row.items())[:5])  # First 5 columns
        is_valid = ask_ollama_is_valid_row(row_sample)
        print(f"Ollama edge case check: {row_sample} -> {'VALID' if is_valid else 'SKIP'}")
        return not is_valid

    return False


def process_vapt_row(row, idx, dsn, rc_lower):
    """Process a SAST_DAST format row - preserves all SAST_DAST columns for display"""
    def get_val(patterns):
        for p in patterns:
            p_lower = p.lower().replace(" ", "").replace("_", "")
            for col_lower, col in rc_lower.items():
                col_normalized = col_lower.replace(" ", "").replace("_", "")
                if p_lower == col_normalized or p_lower in col_normalized:
                    val = str(row.get(col, "")).strip()
                    if val and val.lower() not in ["", "nan", "none", "na", "null"]:
                        return val
        return ""

    rec = {"UploadBatch": dsn, "SourceFormat": "SAST_DAST"}

    # Issue key (primary ID)
    issue_key = get_val(["Issue key", "IssueKey", "Issue_key", "ID"])
    rec["IssueID"] = issue_key if issue_key else f"SAST_DAST-{idx}"
    rec["DisplayID"] = rec["IssueID"]
    rec["issue_key"] = issue_key  # Preserve original column name

    # Summary
    summary = get_val(["Summary", "Title", "Issue", "Description"])
    rec["Name"] = summary
    rec["Summary"] = summary
    rec["Description"] = summary

    # Application Name
    app_name = get_val(["Application Name", "ApplicationName", "Application", "App"])
    rec["AffectedAsset"] = app_name
    rec["ApplicationName"] = app_name
    rec["AssetType"] = "Application"

    # Criticality Status - check "Criticality" column first as it's common in Excel
    criticality = get_val(["Criticality", "Criticality Status", "CriticalityStatus", "Severity", "Priority"])
    rec["CriticalityStatus"] = criticality
    rec["Criticality"] = criticality

    # Severity from Criticality - preserve the actual value
    sev_map = {"critical": "Critical", "high": "High", "medium": "Medium", "low": "Low",
               "exception": "Medium", "info": "Info"}
    rec["Severity"] = sev_map.get(criticality.lower(), criticality) if criticality else "Medium"

    # Compliant/Non-Compliant = Status
    compliant = get_val(["Compliant/Non-compliant", "Compliant", "Status", "Compliance"])
    rec["Compliant_NonCompliant"] = compliant  # Preserve original
    if compliant:
        if "non" in compliant.lower() or "open" in compliant.lower():
            rec["Status"] = "Open"
        else:
            rec["Status"] = "Resolved"
    else:
        rec["Status"] = "Open"

    # Reported on (date)
    reported_on = get_val(["Reported on", "ReportedOn", "Reported_on", "Created", "Date"])
    rec["DiscoveredDate"] = reported_on
    rec["ReportedOn"] = reported_on  # Preserve original

    # Expected Timeline (due date)
    expected_timeline = get_val(["Expected Timeline", "ExpectedTimeline", "Due Date", "DueDate", "Deadline"])
    rec["DueDate"] = expected_timeline
    rec["ExpectedTimeline"] = expected_timeline  # Preserve original

    # Ageing (days open)
    ageing = get_val(["Ageing", "Age", "Days Open"])
    rec["Ageing"] = ageing

    # Assignee
    assignee = get_val(["Assignee", "Assigned To", "AssignedTo", "Owner"])
    rec["Assignee"] = assignee
    rec["AssignedTo"] = assignee

    # Multiple Assignee
    multiple_assignee = get_val(["Multiple Assignee", "MultipleAssignee", "Additional Assignees"])
    rec["MultipleAssignee"] = multiple_assignee
    if not rec["AssignedTo"]:
        rec["AssignedTo"] = multiple_assignee

    # Application Owner / Lob Head
    app_owner = get_val(["Application Owner", "ApplicationOwner", "App Owner", "Lob Head", "LobHead", "LOB Head"])
    rec["ApplicationOwner"] = app_owner
    rec["Department"] = app_owner

    # Category
    rec["Category"] = "SAST_DAST Finding"

    # Generate short vulnerability description (5-7 words)
    rec["VulnDescription"] = generate_short_description(
        summary,
        rec["IssueID"],
        rec["Severity"],
        "Application",
        summary
    )

    # Auto-assign POD owner if not assigned
    if not rec["AssignedTo"] or rec["AssignedTo"] in ["", "NA", "Unassigned"]:
        auto_owner = get_pod_owner(app_name, app_owner, "")
        if auto_owner:
            rec["AssignedTo"] = auto_owner

    lob = get_val(["LOB", "Line of Business", "BusinessUnit"])
    rec["LOB"] = lob if lob else "SAST_DAST"

    return rec


def process_vapt_row_new(row, idx, dsn, rc_lower):
    def get_val(patterns):
        for p in patterns:
            p_lower = p.lower().replace(" ", "").replace("_", "")
            for col_lower, col in rc_lower.items():
                col_normalized = col_lower.replace(" ", "").replace("_", "")
                if p_lower == col_normalized or p_lower in col_normalized:
                    val = str(row.get(col, "")).strip()
                    if val and val.lower() not in ["", "nan", "none", "na", "null"]:
                        return val
        return ""

    rec = {"UploadBatch": dsn, "SourceFormat": "VAPT"}

    uuid = get_val(["UUID", "ID"])
    rec["IssueID"] = uuid if uuid else f"VAPT-{idx}"
    rec["UUID"] = uuid

    vuln_name = get_val(["Vulnerability name", "VulnerabilityName", "Vuln Name"])
    rec["DisplayID"] = vuln_name if vuln_name else rec["IssueID"]
    rec["Vulnerability name"] = vuln_name

    rec["IP"] = get_val(["IP", "IP Address", "IPAddress"])
    rec["Hostname"] = get_val(["Hostname", "Host"])
    rec["Port"] = get_val(["Port"])
    rec["Protocol"] = get_val(["Protocol"])

    rec["Vulnerability description"] = get_val(["Vulnerability description", "VulnerabilityDescription", "Description"])
    rec["Solution"] = get_val(["Solution", "Remediation", "Fix"])
    rec["Vulnerability Path"] = get_val(["Vulnerability Path", "VulnerabilityPath", "Path"])
    rec["Vulnerability ID"] = get_val(["Vulnerability ID", "VulnerabilityID", "Vuln ID"])
    rec["Vulnerability family"] = get_val(["Vulnerability family", "VulnerabilityFamily", "Family", "Category"])
    rec["CVE Number"] = get_val(["CVE Number", "CVENumber", "CVE"])

    risk_factor = get_val(["Risk Factor", "RiskFactor", "Risk", "Severity"])
    severity_map = {"critical": "Critical", "high": "High", "medium": "Medium", "low": "Low", "info": "Info"}
    rec["Severity"] = severity_map.get(risk_factor.lower(), risk_factor) if risk_factor else "Medium"
    rec["Risk Factor"] = risk_factor
    rec["RiskFactor"] = risk_factor

    rec["vprScore"] = get_val(["vprScore", "VPRScore", "VPR Score", "VPR"])
    rec["Priority"] = get_val(["Priority"])

    status = get_val(["Vulnerability Status", "VulnerabilityStatus", "Status"])
    rec["Status"] = status if status else "Open"
    rec["Vulnerability Status"] = status if status else "Open"

    rec["Application Owner"] = get_val(["Application Owner", "ApplicationOwner", "Owner", "Assigned To", "AssignedTo"])
    rec["AssignedTo"] = rec["Application Owner"]
    rec["Application Name"] = get_val(["Application Name", "ApplicationName", "App Name", "Application"])
    rec["LOB Name"] = get_val(["LOB Name", "LOBName", "LOB"])
    rec["Repo Name"] = get_val(["Repo Name", "RepoName", "Repository"])

    rec["firstSeen"] = get_val(["firstSeen", "FirstSeen", "First Seen"])
    rec["lastSeen"] = get_val(["lastSeen", "LastSeen", "Last Seen"])
    rec["DiscoveredDate"] = rec["firstSeen"] if rec["firstSeen"] else get_val(["vulnPubDate", "VulnPubDate"])
    rec["DueDate"] = ""

    rec["Vulnerability Type"] = get_val(["Vulnerability Type", "VulnerabilityType", "Type"])
    rec["Internet/Non-Intern"] = get_val(["Internet/Non-Intern", "InternetExposed", "Internet"])
    rec["Quarter"] = get_val(["Quarter"])

    desc = generate_short_description(vuln_name, rec.get("CVENumber", ""), rec["Severity"], "", "")
    rec["Description"] = desc

    if not rec["AssignedTo"] or rec["AssignedTo"] in ["", "NA", "Unassigned"]:
        auto_owner = get_pod_owner(rec.get("ApplicationName", ""), rec.get("LOBName", ""), rec.get("Application Owner", ""))
        if auto_owner:
            rec["AssignedTo"] = auto_owner

    rec["LOB"] = rec.get("LOBName", "VAPT")

    return rec


def process_cspm_row(row, idx, dsn, rc_lower):
    """Process a CSPM format row"""
    def get_val(patterns):
        for p in patterns:
            p_lower = p.lower().replace(" ", "").replace("_", "")
            for col_lower, col in rc_lower.items():
                col_normalized = col_lower.replace(" ", "").replace("_", "")
                if p_lower == col_normalized or p_lower in col_normalized:
                    val = str(row.get(col, "")).strip()
                    if val and val.lower() not in ["", "nan", "none", "na", "null"]:
                        return val
        return ""

    rec = {"UploadBatch": dsn, "SourceFormat": "CSPM"}

    # IDs - preserve original column names for UI
    resource_id = get_val(["resource_id", "ResourceID", "Resource ID"])
    finding_type_id = get_val(["finding_type_id", "FindingTypeID", "Finding Type ID"])
    rec["IssueID"] = finding_type_id if finding_type_id else f"CSPM-{idx}"
    rec["DisplayID"] = rec["IssueID"]
    rec["resource_id"] = resource_id
    rec["finding_type_id"] = finding_type_id

    # Finding Name = Name
    finding_name = get_val(["finding_name", "FindingName", "Finding Name", "Finding"])
    rec["Name"] = finding_name
    rec["finding_name"] = finding_name
    rec["Description"] = finding_name[:100] + "..." if len(finding_name) > 100 else finding_name

    # Resource = AffectedAsset
    resource_name = get_val(["resource_name", "ResourceName", "Resource Name", "Resource"])
    rec["AffectedAsset"] = resource_name if resource_name else resource_id
    rec["resource_name"] = resource_name
    rec["AssetID"] = resource_id

    # Resource Type
    resource_type = get_val(["resource_type", "ResourceType", "Resource Type", "Type"])
    rec["resource_type"] = resource_type
    rec["AssetType"] = resource_type if resource_type else "Cloud Resource"

    # Cloud Provider & Account - preserve as account_name and account_id
    cloud_provider = get_val(["cloud_provider", "CloudProvider", "Cloud Provider", "Provider"])
    account_name = get_val(["account_name", "AccountName", "Account Name", "Account"])
    account_id = get_val(["account_id", "AccountID", "Account ID", "AccountId"])
    rec["CloudProvider"] = cloud_provider
    rec["CloudPlatform"] = account_name
    rec["account_name"] = account_name
    rec["account_id"] = account_id

    # Severity - preserve actual value from Excel
    severity = get_val(["severity", "Severity", "Risk", "RiskLevel"])
    sev_map = {"critical": "Critical", "high": "High", "medium": "Medium", "low": "Low", "info": "Info"}
    rec["Severity"] = sev_map.get(severity.lower(), severity) if severity else "Medium"
    rec["OriginalSeverity"] = severity

    # Risk Score
    risk_score = get_val(["risk_score", "RiskScore", "Risk Score", "Score"])
    rec["Score"] = risk_score

    # Impact - preserve as 'impact' for UI
    impact = get_val(["impact", "Impact", "Business Impact"])
    rec["Impact"] = impact
    rec["impact"] = impact

    # Compliance Tags - preserve as 'compliance_tags' for UI
    compliance_tags = get_val(["compliance_tags", "ComplianceTags", "Compliance Tags", "Compliance"])
    rec["Tags"] = compliance_tags
    rec["compliance_tags"] = compliance_tags
    rec["Category"] = "CSPM Finding"

    # Generate short vulnerability description (5-7 words)
    rec["VulnDescription"] = generate_short_description(
        finding_name,
        rec["IssueID"],
        rec["Severity"],
        resource_type,
        finding_name
    )

    # Remediation
    remediation = get_val(["remediation", "Remediation", "Remediation Steps", "Fix"])
    remediation_region = get_val(["remediation_region", "RemediationRegion", "Region"])
    rec["RecommendedAction"] = remediation if remediation else "Review cloud configuration"
    rec["LocationPath"] = remediation_region

    # Status (CSPM usually shows current state)
    rec["Status"] = "Open"

    # Dates
    rec["DiscoveredDate"] = datetime.now().strftime("%Y-%m-%d")
    sev = rec["Severity"]
    days = 7 if sev == "Critical" else (30 if sev == "High" else 60)
    rec["DueDate"] = (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d")

    # LOB filter
    lob = get_val(["LOB", "Line of Business", "BusinessUnit"])
    rec["LOB"] = lob
    lob_value = lob.lower().strip() if lob else ""
    if lob_value and lob_value not in ALLOWED_LOB and "wynk" not in lob_value:
        return None

    # Auto-assign for CSPM based on account_id and account_name
    assigned_owner = get_cspm_pod_owner(account_id, account_name, resource_id, resource_name)
    rec["AssignedTo"] = assigned_owner

    return rec


def classify_container_subtype(rec: dict) -> str:
    """
    Classifies a container vulnerability record into exactly one of:
    "Zero day VA", "Wiz CLI Integration", "Compliance VA", "Quarterly VA", "Unclassified"
    """
    tags = str(rec.get("Tags") or "").lower()
    projects = str(rec.get("Projects") or "").lower()
    det_method = str(rec.get("DetectionMethod") or "").lower()
    name = str(rec.get("Name") or "").lower()
    det_name = str(rec.get("DetailedName") or "").lower()
    remediation = str(rec.get("Remediation") or "").lower()
    severity = str(rec.get("Severity") or "Medium").lower()
    first_detected = str(rec.get("DiscoveredDate") or rec.get("FirstDetected") or "")

    # 1. Zero day VA
    fixed_version = str(rec.get("FixedVersion") or "").strip().lower()
    missing_fixed_version = fixed_version in ["", "null", "none", "nan", "unmatched", "n/a", "na"]
    
    is_high_severity = severity in ["high", "critical"]
    is_recent = False
    
    if first_detected:
        try:
            from datetime import datetime, timezone
            if "T" in first_detected:
                dt_obj = datetime.strptime(first_detected.split("T")[0], "%Y-%m-%d")
            elif " " in first_detected:
                dt_obj = datetime.strptime(first_detected.split(" ")[0], "%Y-%m-%d")
            else:
                parts = first_detected.replace("/", "-").split("-")
                if len(parts) == 3:
                    if len(parts[0]) == 4:
                        dt_obj = datetime.strptime(first_detected.replace("/", "-"), "%Y-%m-%d")
                    else:
                        dt_obj = datetime.strptime(first_detected.replace("/", "-"), "%d-%m-%Y")
                else:
                    raise ValueError
            
            dt_obj = dt_obj.replace(tzinfo=timezone.utc)
            now_utc = datetime.now(timezone.utc)
            if (now_utc - dt_obj).days <= 7:
                is_recent = True
        except Exception:
            pass

    if missing_fixed_version or (is_high_severity and is_recent):
        return "Zero day VA"

    # 2. Wiz CLI Integration
    wiz_cli_indicators = ["wizcli", "wiz-cli", "ci/cd", "pipeline", "github-actions", "gitlab-ci", "jenkins"]
    if "filepath" in det_method or any(ind in tags for ind in wiz_cli_indicators) or any(ind in projects for ind in wiz_cli_indicators):
        return "Wiz CLI Integration"

    # 3. Compliance VA
    compliance_keywords = ["compliance", "cis", "pci", "nist", "soc2", "gdpr", "hipaa", "baseline", "policy", "regulatory"]
    compliance_string = f"{tags} {name} {det_name} {det_method} {remediation} {projects}"
    if any(kw in compliance_string for kw in compliance_keywords):
        return "Compliance VA"

    # 4. Quarterly VA
    if first_detected:
        try:
            from datetime import datetime, timezone
            if "T" in first_detected:
                dt_obj = datetime.strptime(first_detected.split("T")[0], "%Y-%m-%d")
            elif " " in first_detected:
                dt_obj = datetime.strptime(first_detected.split(" ")[0], "%Y-%m-%d")
            else:
                parts = first_detected.replace("/", "-").split("-")
                if len(parts) == 3:
                    if len(parts[0]) == 4:
                        dt_obj = datetime.strptime(first_detected.replace("/", "-"), "%Y-%m-%d")
                    else:
                        dt_obj = datetime.strptime(first_detected.replace("/", "-"), "%d-%m-%Y")
                else:
                    raise ValueError
            
            dt_obj = dt_obj.replace(tzinfo=timezone.utc)
            now_utc = datetime.now(timezone.utc)
            if (now_utc - dt_obj).days > 90:
                return "Quarterly VA"
        except Exception:
            pass

    # 5. Unclassified
    return "Unclassified"


def process_container_row(row, idx, dsn, rc_lower):
    """Process a Container/Container Image format row"""
    def get_val(patterns):
        for p in patterns:
            p_lower = p.lower().replace(" ", "").replace("_", "")
            for col_lower, col in rc_lower.items():
                col_normalized = col_lower.replace(" ", "").replace("_", "")
                if p_lower == col_normalized or p_lower in col_normalized:
                    val = str(row.get(col, "")).strip()
                    if val and val.lower() not in ["", "nan", "none", "na", "null"]:
                        return val
        return ""

    rec = {"UploadBatch": dsn, "SourceFormat": "CONTAINER"}

    # Copy all original columns
    for k, v in row.items():
        rec[k] = str(v).strip() if v is not None else ""

    # IssueID and DisplayID
    issue_id = get_val(["ID", "IssueID", "VulnID", "CVE", "VulnerabilityID"])
    rec["IssueID"] = issue_id if issue_id else f"VULN-{idx}"

    display_id = get_val(["ID", "DisplayID", "CVE", "VulnerabilityID", "Name"])
    if display_id and display_id.upper().startswith("CVE"):
        rec["DisplayID"] = display_id
    elif issue_id and issue_id.upper().startswith("CVE"):
        rec["DisplayID"] = issue_id
    elif display_id:
        rec["DisplayID"] = display_id
    else:
        rec["DisplayID"] = rec["IssueID"]

    # Severity - preserve actual value from Excel
    sev = get_val(["Severity", "CVSSSeverity", "VendorSeverity", "NvdSeverity", "Risk", "RiskLevel"])
    rec["OriginalSeverity"] = sev
    if sev:
        sev_lower = sev.lower().strip()
        if sev_lower == "critical" or "critical" in sev_lower:
            rec["Severity"] = "Critical"
        elif sev_lower == "high" or "high" in sev_lower:
            rec["Severity"] = "High"
        elif sev_lower == "medium" or "medium" in sev_lower or "moderate" in sev_lower:
            rec["Severity"] = "Medium"
        elif sev_lower == "low" or "low" in sev_lower:
            rec["Severity"] = "Low"
        elif sev_lower == "info" or "info" in sev_lower:
            rec["Severity"] = "Info"
        else:
            rec["Severity"] = sev
    else:
        rec["Severity"] = "Medium"

    # Status
    status = get_val(["Status", "State", "FindingStatus"])
    rec["Status"] = status if status else "Open"

    # Category
    category = get_val(["Category", "Type", "VulnType", "AssetType"])
    rec["Category"] = category if category else "CONTAINER_IMAGE"

    # Name and DetailedName
    rec["Name"] = get_val(["Name", "VulnerabilityName", "Title", "Summary"])
    rec["DetailedName"] = get_val(["DetailedName", "DetailName", "FullName"])

    # AffectedAsset - IMPORTANT: map from AssetName
    affected_asset = get_val(["AffectedAsset", "AssetName", "Asset", "Host", "Hostname", "Target", "Resource"])
    rec["AffectedAsset"] = affected_asset
    rec["AssetID"] = get_val(["AssetID", "AssetId", "ResourceID"])
    rec["AssetType"] = get_val(["AssetType", "ResourceType", "TargetType"])

    # DiscoveredDate and DueDate - IMPORTANT
    discovered = get_val(["DiscoveredDate", "FirstDetected", "DetectedDate", "FoundDate", "CreatedDate"])
    rec["DiscoveredDate"] = discovered
    rec["FirstDetected"] = get_val(["FirstDetected", "FirstDetec", "FirstSeen"])

    due = get_val(["DueDate", "Due", "Deadline", "TargetDate"])
    if due:
        rec["DueDate"] = due
    elif discovered:
        try:
            dt = pd.to_datetime(discovered, errors='coerce')
            if pd.notna(dt):
                days = 7 if rec["Severity"] == "Critical" else (30 if rec["Severity"] == "High" else 60)
                rec["DueDate"] = (dt + pd.Timedelta(days=days)).strftime("%Y-%m-%d")
        except:
            rec["DueDate"] = ""
    else:
        rec["DueDate"] = ""

    # Subscription info - IMPORTANT for auto-assignment
    subscription_name = get_val(["SubscriptionName", "SubName", "SubscriptionExternalId"])
    subscription_id = get_val(["SubscriptionId", "SubscriptionID", "SubID", "SubscriptionExternalId"])
    rec["SubscriptionName"] = subscription_name
    rec["SubscriptionId"] = subscription_id

    # Department and AssignedTo
    rec["Department"] = get_val(["Department", "AssignedTeam", "Team", "LOB"])
    assigned_to = get_val(["AssignedTo", "Assignee", "Owner"])
    rec["AssignedTo"] = assigned_to

    # Auto-assign POD owner based on subscription name/ID
    if not rec["AssignedTo"] or rec["AssignedTo"] in ["", "NA", "Unassigned"]:
        auto_owner = get_pod_owner(rec.get("SubscriptionName", ""), rec.get("SubscriptionId", ""), rec.get("AffectedAsset", ""), rec.get("Projects", ""))
        if auto_owner:
            rec["AssignedTo"] = auto_owner

    # LOB
    lob = get_val(["LOB", "LineOfBusiness", "BusinessUnit"])
    rec["LOB"] = lob

    # Remediation and Links
    rec["RecommendedAction"] = get_val(["RecommendedAction", "Remediation", "Resolution", "Fix", "Mitigation"])
    rec["ReferenceLinks"] = get_val(["Link", "URL", "WizURL", "Reference", "ReferenceLink"])
    rec["WizURL"] = get_val(["WizURL", "WizLink"])

    # Version info
    rec["Version"] = get_val(["Version", "CurrentVersion", "InstalledVersion"])
    rec["FixedVersion"] = get_val(["FixedVersion", "PatchedVersion", "RemediatedVersion"])
    rec["Score"] = get_val(["Score", "CVSSScore", "CVSS", "CVSSv3"])

    # Additional fields
    rec["CVSSSeverity"] = get_val(["CVSSSeverity", "CVSSSev"])
    rec["VendorSeverity"] = get_val(["VendorSeverity", "VendorSev"])
    rec["NvdSeverity"] = get_val(["NvdSeverity", "NVDSev"])
    rec["HasExploit"] = get_val(["HasExploit", "ExploitAvailable", "Exploitable"])
    rec["HasCisaKev"] = get_val(["HasCisaKev", "HasCisaKnownExploit", "CisaKEV"])
    rec["FindingStatus"] = get_val(["FindingStatus", "FindingStat"])
    rec["LastDetected"] = get_val(["LastDetected", "LastDetec", "LastSeen"])
    rec["ResolvedAt"] = get_val(["ResolvedAt", "ResolvedDate", "FixedDate"])
    rec["Resolution"] = get_val(["Resolution", "ResolutionStatus"])
    rec["LocationPath"] = get_val(["LocationPath", "Location", "Path"])
    rec["Projects"] = get_val(["Projects", "Project", "Application"])
    rec["CloudProvider"] = get_val(["CloudProvider", "Provider", "Cloud"])
    rec["CloudPlatform"] = get_val(["CloudPlatform", "Platform"])
    rec["Namespaces"] = get_val(["Namespaces", "Namespace", "NS"])
    rec["Clusters"] = get_val(["Clusters", "Cluster", "K8sCluster"])
    rec["Tags"] = get_val(["Tags", "Tag", "Labels"])
    rec["DetectionMethod"] = get_val(["DetectionMethod", "Detection", "Method"])

    rec["ContainerSubType"] = classify_container_subtype(rec)


    # Generate short vulnerability description
    rec["Description"] = generate_short_description(
        rec["Name"], rec["DisplayID"], rec["Severity"], rec["AssetType"], rec["DetailedName"]
    )
    rec["VulnDescription"] = rec["Description"]

    # LOB filter - skip non-Wynk
    lob_value = lob.lower().strip() if lob else ""
    if lob_value and lob_value not in ALLOWED_LOB and "wynk" not in lob_value:
        return None

    return rec


def detect_header_row(ws, max_rows=15):
    """Search first max_rows rows to find the header row with most expected columns."""
    best_row = 1
    best_count = 0
    best_cols = []

    for row_idx in range(1, min(max_rows + 1, ws.max_row + 1)):
        row_values = []
        for col_idx in range(1, min(ws.max_column + 1, 50)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                row_values.append(str(cell.value).strip().lower())
            else:
                row_values.append("")

        matched_cols = [v for v in row_values if v and v.replace(" ", "").replace("_", "") in EXPECTED_COLUMNS]

        if len(matched_cols) > best_count:
            best_count = len(matched_cols)
            best_row = row_idx
            best_cols = matched_cols

    return best_row, best_count, best_cols


def is_pivot_table_sheet(ws, header_row):
    """
    Detect if a sheet is a pivot table or summary sheet (not raw data).
    Pivot tables typically have:
    - "Row Labels" / "Column Labels" headers
    - "Count of X", "Sum of X" headers
    - "Grand Total" rows
    - Very few columns (2-4) with mostly numeric data
    """
    header_values_raw = []
    for col_idx in range(1, min(ws.max_column + 1, 20)):
        cell = ws.cell(row=header_row, column=col_idx)
        if cell.value is not None:
            header_values_raw.append(str(cell.value).strip().lower())

    # Check for pivot table indicators
    pivot_score = 0
    for header in header_values_raw:
        for indicator in PIVOT_INDICATORS:
            if indicator in header:
                pivot_score += 10

    # Check for "Row Labels" which is a dead giveaway for Excel pivot tables
    if any("row labels" in h or "rowlabels" in h.replace(" ", "") for h in header_values_raw):
        pivot_score += 50

    # If very few columns (2-4) and one is a count/sum, likely a pivot
    if len(header_values_raw) <= 4:
        pivot_score += 5

    return pivot_score >= 10


def score_sheet(ws, sheet_name):
    """Score a worksheet based on expected columns and data characteristics."""
    score = 0
    details = []
    is_pivot = False

    header_row, col_count, matched_cols = detect_header_row(ws)

    # Check for negative patterns in sheet name
    sheet_name_lower = sheet_name.lower()
    for pattern in NEGATIVE_PATTERNS:
        if pattern in sheet_name_lower:
            score -= 5
            details.append(f"-5 (sheet name contains '{pattern}')")

    # Get header values (both raw and normalized)
    header_values = set()
    header_values_raw = []
    for col_idx in range(1, min(ws.max_column + 1, 50)):
        cell = ws.cell(row=header_row, column=col_idx)
        if cell.value is not None:
            raw_val = str(cell.value).strip().lower()
            header_values_raw.append(raw_val)
            header_values.add(raw_val.replace(" ", "").replace("_", ""))

    # Check if this is a pivot/summary table - heavily penalize
    if is_pivot_table_sheet(ws, header_row):
        score -= 100
        is_pivot = True
        details.append(f"-100 (detected as pivot/summary table)")

    # Check for negative patterns in headers
    for col in header_values:
        for pattern in NEGATIVE_PATTERNS:
            if pattern.replace(" ", "") in col:
                score -= 5
                details.append(f"-5 (header contains '{pattern}')")

    # Score high-value columns (only if not a pivot)
    if not is_pivot:
        for col in HIGH_SCORE_COLS:
            if col in header_values:
                score += 3
                details.append(f"+3 ({col})")

        # Score medium-value columns
        for col in MED_SCORE_COLS:
            if col in header_values:
                score += 2
                details.append(f"+2 ({col})")

        # Score other expected columns
        other_cols = EXPECTED_COLUMNS - HIGH_SCORE_COLS - MED_SCORE_COLS
        for col in other_cols:
            if col in header_values:
                score += 1
                details.append(f"+1 ({col})")

        all_format_cols = VAPT_COLUMNS | SAST_DAST_COLUMNS | CSPM_COLUMNS | CONTAINER_COLUMNS
        for col in all_format_cols:
            col_normalized = col.replace(" ", "").replace("_", "")
            if col_normalized in header_values or col in [h for h in header_values_raw]:
                score += 2
                details.append(f"+2 (format-specific: {col})")

    # Column count scoring - more columns = more likely raw data
    num_cols = len([h for h in header_values_raw if h])
    if num_cols >= 10:
        score += 10
        details.append(f"+10 (>= 10 columns: {num_cols})")
    elif num_cols >= 6:
        score += 5
        details.append(f"+5 (>= 6 columns: {num_cols})")
    elif num_cols <= 3:
        score -= 10
        details.append(f"-10 (<= 3 columns: {num_cols}, likely summary)")

    # Row count scoring
    data_rows = ws.max_row - header_row
    if data_rows > 100:
        score += 5
        details.append(f"+5 (>100 rows: {data_rows})")
    elif data_rows > 50:
        score += 3
        details.append(f"+3 (>50 rows: {data_rows})")
    elif data_rows < 20:
        score -= 3
        details.append(f"-3 (<20 rows: {data_rows})")

    return {
        "sheet_name": sheet_name,
        "score": score,
        "header_row": header_row,
        "data_rows": data_rows,
        "matched_columns": col_count,
        "num_columns": num_cols,
        "is_pivot": is_pivot,
        "details": details
    }


def find_best_worksheet(file_bytes):
    """Find the best worksheet containing vulnerability data."""
    print("Scanning workbook for vulnerability data...")

    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_scores = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Skip empty sheets
        if ws.max_row is None or ws.max_row < 2:
            print(f"  Skipping '{sheet_name}': empty or single row")
            continue

        result = score_sheet(ws, sheet_name)
        sheet_scores.append(result)

        pivot_flag = " [PIVOT/SUMMARY]" if result.get("is_pivot") else ""
        print(f"  Checking worksheet: {sheet_name}{pivot_flag}")
        print(f"    Score: {result['score']} | Header row: {result['header_row']} | Data rows: {result['data_rows']} | Columns: {result.get('num_columns', 'N/A')}")

    wb.close()

    if not sheet_scores:
        return None, []

    # Sort by score descending (pivot tables will have negative scores)
    sheet_scores.sort(key=lambda x: x["score"], reverse=True)
    best = sheet_scores[0]

    print(f"  Best worksheet: {best['sheet_name']} (score: {best['score']})")

    # If best sheet is still a pivot/summary, try to find a non-pivot sheet
    if best.get("is_pivot") and len(sheet_scores) > 1:
        non_pivot_sheets = [s for s in sheet_scores if not s.get("is_pivot")]
        if non_pivot_sheets:
            best = non_pivot_sheets[0]
            print(f"  Switched to non-pivot sheet: {best['sheet_name']} (score: {best['score']})")

    # Confidence check: if best has fewer than 5 matched columns, flag for manual selection
    # But don't flag if we have a clear winner with many columns
    if best.get("num_columns", 0) < 5 and len(sheet_scores) > 1:
        print(f"  Low confidence: only {best.get('num_columns', 0)} columns found")
        # Filter out pivot tables from the selection list
        selectable_sheets = [s for s in sheet_scores if not s.get("is_pivot")]
        if selectable_sheets:
            return None, selectable_sheets
        return None, sheet_scores

    return best, sheet_scores


def read_selected_sheet(file_bytes, sheet_name, header_row):
    """Read the selected worksheet starting from the detected header row."""
    print(f"  Reading worksheet '{sheet_name}' from header row {header_row}...")

    df = pd.read_excel(
        BytesIO(file_bytes),
        sheet_name=sheet_name,
        header=header_row - 1  # pandas uses 0-based index
    )

    print(f"  Rows loaded: {len(df)}")
    return df

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
dbf = "xtelify_db.json"


schema_cache = {}

def remove_duplicates(data):
    """Remove exact duplicate records where EVERY column is same"""
    seen = set()
    unique = []
    for item in data:
        # Create a unique key from ALL values (sorted to ensure consistency)
        key_parts = []
        for k in sorted(item.keys()):
            key_parts.append(f"{k}:{item.get(k, '')}")
        key = "|".join(key_parts)

        if key not in seen:
            seen.add(key)
            unique.append(item)

    removed = len(data) - len(unique)
    if removed > 0:
        print(f"[DB] Removed {removed} exact duplicate records")
    return unique

DB_CACHE_MAX_RECORDS = 10000
MONGO_HEALTH_TTL = 5
_mongo_last_health_check = 0
_mongo_is_available = None


def get_file_hash():
    """Legacy cache key helper kept for compatibility."""
    return None


def _is_mongo_available(force_check=False):
    global _mongo_last_health_check, _mongo_is_available

    now = time.time()
    if not force_check and _mongo_is_available is not None and (now - _mongo_last_health_check) < MONGO_HEALTH_TTL:
        return _mongo_is_available

    try:
        client.admin.command("ping")
        _mongo_is_available = True
    except Exception as e:
        _mongo_is_available = False
        print(f"[DB] MongoDB connection failed: {e}")

    _mongo_last_health_check = now
    return _mongo_is_available


def _ensure_mongo_indexes():
    if not _is_mongo_available(force_check=True):
        print("[DB] Skipping MongoDB index setup because database is unavailable")
        return

    try:
        issues_collection.create_index([("IssueID", 1)])
        issues_collection.create_index([("UploadBatch", 1)])
        issues_collection.create_index([("Severity", 1)])
        issues_collection.create_index([("Status", 1)])
        issues_collection.create_index([("AssignedTo", 1)])
        issues_collection.create_index([("UploadedAt", -1)])
        issues_collection.create_index([("FileHash", 1)])
        upload_history_collection.create_index([("UploadedAt", -1)])
        upload_history_collection.create_index([("UploadBatch", 1)])
        
        # AI Remediation Cache Indexes
        ai_remediation_cache_collection.create_index([("IssueID", 1), ("UploadBatch", 1), ("SourceFormat", 1)])
        
        print("[DB] MongoDB indexes ensured for all collections")
    except Exception as e:
        print(f"[DB] Failed to create MongoDB indexes: {e}")


def _ensure_uploaded_at(rec):
    if "UploadedAt" not in rec or rec.get("UploadedAt") in [None, ""]:
        rec["UploadedAt"] = datetime.now(timezone.utc)
    return rec


def _prepare_records_for_write(records, ensure_uploaded_at=True):
    if not isinstance(records, list):
        return []

    unique_data = remove_duplicates(records)
    prepared = []

    for item in unique_data:
        if not isinstance(item, dict):
            continue
        rec = dict(item)
        if ensure_uploaded_at:
            _ensure_uploaded_at(rec)
        prepared.append(rec)

    return prepared


def _auto_correct_assigned_to(rec):
    current_owner = rec.get("AssignedTo", "Unassigned")
    if current_owner not in ["", "NA", "Unassigned", "None"]:
        return None

    fields = [
        rec.get("account_name"), rec.get("account_id"),
        rec.get("SubscriptionName"), rec.get("SubscriptionId"),
        rec.get("Projects"), rec.get("ApplicationName"),
        rec.get("AffectedAsset"), rec.get("resource_name"),
        rec.get("resource_id"), rec.get("LOB Name"),
        rec.get("LOBName")
    ]

    if rec.get("SourceFormat") == "CSPM":
        auto_owner = get_cspm_pod_owner(*fields)
        if auto_owner != "Unassigned":
            rec["AssignedTo"] = auto_owner
            return auto_owner
    else:
        auto_owner = get_pod_owner(*fields)
        if auto_owner:
            rec["AssignedTo"] = auto_owner
            return auto_owner

    return None


def _record_signature(rec):
    normalized = {}
    for key in sorted(rec.keys()):
        if key in ["_id", "UploadedAt"]:
            continue
        value = rec.get(key)
        if isinstance(value, datetime):
            normalized[key] = value.isoformat()
        else:
            normalized[key] = value

    payload = json.dumps(normalized, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def ldb(use_cache=True):
    global _db_cache, _db_cache_time, _db_cache_hash

    if use_cache and _db_cache is not None and (time.time() - _db_cache_time) < CACHE_TTL:
        return _db_cache

    if not _is_mongo_available():
        return []

    try:
        data = []
        owner_updates = []

        for rec in issues_collection.find({}):
            if not isinstance(rec, dict):
                continue

            doc_id = rec.get("_id")
            updated_owner = _auto_correct_assigned_to(rec)
            if updated_owner and doc_id is not None:
                owner_updates.append(UpdateOne({"_id": doc_id}, {"$set": {"AssignedTo": updated_owner}}))

            rec.pop("_id", None)
            data.append(rec)

        if owner_updates:
            issues_collection.bulk_write(owner_updates, ordered=False)

        if use_cache and len(data) <= DB_CACHE_MAX_RECORDS:
            _db_cache = data
            _db_cache_time = time.time()
            _db_cache_hash = "mongodb"
        else:
            _db_cache = None
            _db_cache_time = 0
            _db_cache_hash = None

        print(f"[DB] Loaded {len(data)} records from MongoDB")
        return data
    except Exception as mongo_error:
        print(f"[DB] MongoDB load failed: {mongo_error}")
        return []


def sdb(d):
    global _db_cache, _db_cache_time, _db_cache_hash

    if not _is_mongo_available():
        print("[DB] MongoDB save skipped: database unavailable")
        raise RuntimeError("MongoDB unavailable")

    try:
        prepared = _prepare_records_for_write(d, ensure_uploaded_at=True)
        issues_collection.delete_many({})
        if prepared:
            issues_collection.insert_many(prepared, ordered=False)

        if len(prepared) <= DB_CACHE_MAX_RECORDS:
            _db_cache = prepared
            _db_cache_time = time.time()
            _db_cache_hash = "mongodb"
        else:
            _db_cache = None
            _db_cache_time = 0
            _db_cache_hash = None

        print(f"[DB] Saved {len(prepared)} records to MongoDB")
    except Exception as mongo_error:
        print(f"[DB] MongoDB save failed: {mongo_error}")
        raise


def clear_cache():
    """Clear the database cache"""
    global _db_cache, _db_cache_time, _db_cache_hash
    _db_cache = None
    _db_cache_time = 0
    _db_cache_hash = None


def upload_batch_exists(upload_batch):
    if not upload_batch:
        return False

    if not _is_mongo_available():
        raise RuntimeError("MongoDB unavailable")

    try:
        return issues_collection.find_one({"UploadBatch": upload_batch}, {"_id": 1}) is not None
    except Exception as e:
        print(f"[DB] UploadBatch check failed: {e}")
        raise


def check_duplicate_upload(file_hash):
    if not file_hash:
        return {"duplicate": False}

    if not _is_mongo_available():
        return {"duplicate": False}

    try:
        # Find the most recent upload with this hash
        most_recent = issues_collection.find_one(
            {"FileHash": file_hash},
            sort=[("UploadedAt", -1)]
        )
        if not most_recent:
            return {"duplicate": False}
            
        prev_date = most_recent.get("UploadedAt")
        if not prev_date:
            return {"duplicate": True, "previous_upload_date": "Unknown", "uploaded_today": False}
            
        if prev_date.tzinfo is None:
            prev_date = prev_date.replace(tzinfo=timezone.utc)
            
        local_now = datetime.now()
        local_prev = prev_date.astimezone()
        
        uploaded_today = (local_prev.date() == local_now.date())
        
        formatted_date = local_prev.strftime("%d %B %Y")
        if formatted_date.startswith("0"):
            formatted_date = formatted_date[1:]
            
        return {
            "duplicate": True,
            "previous_upload_date": formatted_date,
            "uploaded_today": uploaded_today
        }
    except Exception as e:
        print(f"[DB] FileHash check failed: {e}")
        return {"duplicate": False}


def attach_file_hash(records, file_hash):
    if not file_hash or not isinstance(records, list):
        return records

    for rec in records:
        if isinstance(rec, dict):
            rec["FileHash"] = file_hash

    return records


def log_upload_history(upload_batch, file_name, source_format, record_count):
    if not _is_mongo_available():
        return
    try:
        now = datetime.now(timezone.utc)
        record = {
            "UploadBatch": upload_batch,
            "FileName": file_name,
            "SourceFormat": source_format,
            "RecordCount": record_count,
            "UploadedAt": now
        }
        upload_history_collection.insert_one(record)
    except Exception as e:
        print(f"[DB] Error logging upload history: {e}")

def insert_records(records, skip_existing_check=False):
    if not records:
        return 0

    if not _is_mongo_available():
        print("[DB] MongoDB insert skipped: database unavailable")
        raise RuntimeError("MongoDB unavailable")

    try:
        prepared = _prepare_records_for_write(records, ensure_uploaded_at=True)
        if not prepared:
            return 0

        to_insert = []
        if skip_existing_check:
            to_insert = prepared
        else:
            for rec in prepared:
                duplicate_query = {k: v for k, v in rec.items() if k != "UploadedAt"}
                if not duplicate_query:
                    to_insert.append(rec)
                    continue

                existing = issues_collection.find_one(duplicate_query, {"_id": 1})
                if existing is None:
                    to_insert.append(rec)

        if to_insert:
            issues_collection.insert_many(to_insert, ordered=False)

        clear_cache()
        skipped = len(prepared) - len(to_insert)
        print(f"[DB] Inserted {len(to_insert)} records into MongoDB (skipped {skipped} duplicates)")
        return len(to_insert)
    except Exception as e:
        print(f"[DB] MongoDB insert failed: {e}")
        raise


def delete_by_upload_batch(upload_batch):
    if not upload_batch:
        return 0

    if not _is_mongo_available():
        print("[DB] MongoDB delete skipped: database unavailable")
        raise RuntimeError("MongoDB unavailable")

    try:
        result = issues_collection.delete_many({"UploadBatch": upload_batch})
        clear_cache()
        print(f"[DB] Deleted {result.deleted_count} records for UploadBatch={upload_batch}")
        return result.deleted_count
    except Exception as e:
        print(f"[DB] MongoDB delete failed: {e}")
        raise


def migrate_json_to_mongodb_once(json_file_path=dbf):
    """One-time safe migration helper from xtelify_db.json to MongoDB."""
    json_abs_path = os.path.abspath(json_file_path)

    if not os.path.exists(json_file_path):
        print(f"[MIGRATE] JSON file not found: {json_abs_path}")
        return {"migrated": 0, "skipped": 0, "source": json_abs_path}

    if not _is_mongo_available(force_check=True):
        print("[MIGRATE] MongoDB unavailable, migration not executed")
        return {"migrated": 0, "skipped": 0, "source": json_abs_path, "error": "MongoDB unavailable"}

    try:
        with open(json_file_path, "r", encoding="utf-8") as f:
            json_records = json.load(f)
    except Exception as e:
        print(f"[MIGRATE] Failed reading JSON file: {e}")
        return {"migrated": 0, "skipped": 0, "source": json_abs_path, "error": str(e)}

    if not isinstance(json_records, list):
        print(f"[MIGRATE] Invalid JSON structure in {json_abs_path}: expected list")
        return {"migrated": 0, "skipped": 0, "source": json_abs_path, "error": "Invalid JSON structure"}

    prepared = _prepare_records_for_write(json_records, ensure_uploaded_at=True)

    try:
        existing_signatures = set()
        for rec in issues_collection.find({}, {"_id": 0}):
            existing_signatures.add(_record_signature(rec))

        to_insert = []
        skipped = 0

        for rec in prepared:
            signature = _record_signature(rec)
            if signature in existing_signatures:
                skipped += 1
                continue

            existing_signatures.add(signature)
            to_insert.append(rec)

        if to_insert:
            issues_collection.insert_many(to_insert, ordered=False)

        clear_cache()

        migrated_count = len(to_insert)
        print(
            f"[MIGRATE] Migrated {migrated_count} records from {json_abs_path} "
            f"(skipped {skipped} exact duplicates). JSON file kept unchanged."
        )
        return {"migrated": migrated_count, "skipped": skipped, "source": json_abs_path}
    except Exception as e:
        print(f"[MIGRATE] Migration failed: {e}")
        return {"migrated": 0, "skipped": 0, "source": json_abs_path, "error": str(e)}


_ensure_mongo_indexes()


def _build_db_query(search=None, search_field=None, severity=None, status=None, assigned_to=None, source_format=None, upload_batch=None, date_from=None, date_to=None, is_advanced_search=None, container_sub_types=None):
    query = {}
    
    if source_format and source_format != "All":
        query["SourceFormat"] = source_format
        
    if container_sub_types and source_format == "CONTAINER":
        if '||' in container_sub_types:
            query["ContainerSubType"] = {"$in": [s.strip() for s in container_sub_types.split("||")]}
        else:
            query["ContainerSubType"] = container_sub_types

    if upload_batch:
        if '||' in upload_batch:
            query["UploadBatch"] = {"$in": [b.strip() for b in upload_batch.split("||")]}
        else:
            query["UploadBatch"] = upload_batch
            
    if is_advanced_search == "true":
        if search:
            s = search.strip()
            regex = {"$regex": s, "$options": "i"}
            
            if search_field and search_field != "All":
                field_map = {
                    "Issue ID": "IssueID",
                    "Finding Name": "finding_name",
                    "Vulnerability Name": "Name",
                    "CVE": "CVE Number",
                    "Account Name": "account_name",
                    "Account ID": "account_id",
                    "Resource Name": "resource_name",
                    "Resource ID": "resource_id",
                    "Assigned To": "AssignedTo",
                    "Hostname": "Hostname",
                    "IP": "IP",
                    "Application": "ApplicationName",
                    "UploadBatch": "UploadBatch"
                }
                db_field = field_map.get(search_field, search_field)
                query[db_field] = regex
            else:
                query["$or"] = [
                    {"DisplayID": regex},
                    {"IssueID": regex},
                    {"AssignedTo": regex},
                    {"RecommendedAction": regex},
                    {"Category": regex},
                    {"Type": regex},
                    {"LOB Name": regex},
                    {"LOBName": regex},
                    {"LOB": regex},
                    {"finding_name": regex},
                    {"FindingName": regex},
                    {"Name": regex},
                    {"VulnDescription": regex},
                    {"Description": regex},
                    {"CVE Number": regex},
                    {"account_name": regex},
                    {"resource_name": regex},
                    {"Hostname": regex},
                    {"IP": regex},
                    {"ApplicationName": regex}
                ]
            
        if severity:
            sev_lower = severity.lower()
            if sev_lower == "critical":
                query["$or"] = [
                    {"Severity": {"$regex": "^(critical|urgent|high)$", "$options": "i"}}, 
                    {"CriticalityStatus": {"$regex": "^(critical|urgent|high)$", "$options": "i"}},
                    {"Criticality": {"$regex": "^(critical|urgent|high)$", "$options": "i"}}
                ]
            else:
                query["Severity"] = {"$regex": f"^{severity}$", "$options": "i"}
                
        if status:
            status_lower = status.lower()
            resolved_keywords = ["resolved", "closed", "fixed", "mitigated", "accepted", "false positive"]
            progress_keywords = ["progress", "pending", "review"]
            
            if status_lower == "resolved":
                query["Status"] = {"$regex": "|".join(resolved_keywords), "$options": "i"}
            elif status_lower == "progress":
                query["Status"] = {"$regex": "|".join(progress_keywords), "$options": "i"}
            elif status_lower == "open":
                query["Status"] = {"$not": {"$regex": "|".join(resolved_keywords + progress_keywords), "$options": "i"}}
            else:
                query["Status"] = status
                
        if assigned_to:
            if assigned_to.lower() == "unassigned":
                query["AssignedTo"] = {"$in": ["", "NA", "Unassigned", None]}
            else:
                if ',' in assigned_to:
                    query["AssignedTo"] = {"$in": [a.strip() for a in assigned_to.split(",")]}
                else:
                    query["AssignedTo"] = assigned_to
                    
        if date_from or date_to:
            date_query = {}
            if date_from:
                date_query["$gte"] = date_from
            if date_to:
                date_query["$lte"] = date_to
            query["UploadedAt"] = date_query
            
    return query

@app.get("/api/db")
async def gd(
    request: Request,
    page: int = 1,
    limit: int = 100,
    search: str = None,
    search_field: str = None,
    severity: str = None,
    status: str = None,
    assigned_to: str = None,
    source_format: str = None,
    upload_batch: str = None,
    date_from: str = None,
    date_to: str = None,
    is_advanced_search: str = None,
    container_sub_types: str = None
):
    if not _is_mongo_available():
        return ORJSONResponse(content={"data": [], "pagination": {"page": page, "limit": limit, "total": 0, "total_pages": 0}})

    query = _build_db_query(
        search=search,
        search_field=search_field,
        severity=severity,
        status=status,
        assigned_to=assigned_to,
        source_format=source_format,
        upload_batch=upload_batch,
        date_from=date_from,
        date_to=date_to,
        is_advanced_search=is_advanced_search,
        container_sub_types=container_sub_types
    )

    try:
        total_records = issues_collection.count_documents(query)
        
        # Determine actual page in bounds
        total_pages = (total_records + limit - 1) // limit if total_records > 0 else 1
        page = min(page, total_pages) if page > 1 else max(1, page)
        
        cursor = issues_collection.find(query).sort("UploadedAt", -1).skip((page - 1) * limit).limit(limit)
        
        records = []
        owner_updates = []
        
        for rec in cursor:
            # Auto-correct assigned to like ldb() did
            doc_id = rec.get("_id")
            updated_owner = _auto_correct_assigned_to(rec)
            if updated_owner and doc_id is not None:
                owner_updates.append(UpdateOne({"_id": doc_id}, {"$set": {"AssignedTo": updated_owner}}))
                rec["AssignedTo"] = updated_owner
                
            rec["_id"] = str(rec["_id"])
            records.append(rec)
            
        if owner_updates:
            # Fire updates asynchronously or wait (doing it synchronously for safety if small batch)
            try:
                issues_collection.bulk_write(owner_updates, ordered=False)
            except Exception as e:
                print(f"[DB Update Error] {e}")

        # Maintain existing unique data logic on the page to prevent duplicate entries if any
        unique_data = remove_duplicates(records)
        
        print(f"[API] /api/db returning {len(unique_data)} records for page {page} out of {total_records} total")
        
        return ORJSONResponse(content={
            "data": unique_data,
            "pagination": {
                "page": page,
                "limit": limit,
                "total": total_records,
                "total_pages": total_pages
            }
        })
    except Exception as e:
        print(f"[API Error] /api/db failed: {e}")
        return ORJSONResponse(status_code=500, content={"error": str(e), "data": [], "pagination": {"total": 0}})


@app.get("/api/container_analytics")
async def container_analytics(
    request: Request,
    assigned_to: str = None
):
    if not _is_mongo_available():
        return ORJSONResponse(content=[])

    query = {"SourceFormat": "CONTAINER"}
    if assigned_to:
        if ',' in assigned_to:
            query["AssignedTo"] = {"$in": [a.strip() for a in assigned_to.split(",")]}
        else:
            query["AssignedTo"] = assigned_to

    try:
        pipeline = [
            {"$match": query},
            {"$group": {
                "_id": {"$ifNull": ["$ContainerSubType", "Unclassified"]},
                "count": {"$sum": 1}
            }}
        ]
        
        results = list(issues_collection.aggregate(pipeline))
        formatted_results = [{"name": r["_id"], "value": r["count"]} for r in results]
        
        # Ensure all types exist even if 0
        all_types = ["Wiz CLI", "Zero-day VA", "Compliance VA", "Quarterly VA", "Unclassified"]
        for t in all_types:
            if not any(r["name"] == t for r in formatted_results):
                formatted_results.append({"name": t, "value": 0})
                
        # Sort by predefined order
        formatted_results.sort(key=lambda x: all_types.index(x["name"]))
        
        return ORJSONResponse(content=formatted_results)
    except Exception as e:
        logger.error(f"Error fetching container analytics: {e}")
        return ORJSONResponse(status_code=500, content={"error": str(e)})

@app.get("/api/db/summary")
async def db_summary(
    request: Request,
    search: str = None,
    search_field: str = None,
    severity: str = None,
    status: str = None,
    assigned_to: str = None,
    source_format: str = None,
    upload_batch: str = None,
    date_from: str = None,
    date_to: str = None,
    is_advanced_search: str = None,
    container_sub_types: str = None
):
    if not _is_mongo_available():
        return ORJSONResponse(content={"total": 0, "status": {"resolved": 0, "open": 0}, "severity": {"critical": 0, "high": 0, "medium": 0, "low": 0}, "cspm": []})

    query = _build_db_query(
        search=search, search_field=search_field, severity=severity, status=status,
        assigned_to=assigned_to, source_format=source_format, upload_batch=upload_batch,
        date_from=date_from, date_to=date_to, is_advanced_search=is_advanced_search,
        container_sub_types=container_sub_types
    )

    try:
        pipeline = [
            {"$match": query},
            {"$facet": {
                "total": [{"$count": "count"}],
                "status": [
                    {"$group": {
                        "_id": {
                            "$cond": [
                                {"$regexMatch": {"input": {"$toLower": "$Status"}, "regex": "resolved|closed|fixed|mitigated|accepted|false positive"}},
                                "resolved",
                                "open"
                            ]
                        },
                        "count": {"$sum": 1}
                    }}
                ],
                "severity_raw": [
                    {"$group": {
                        "_id": {
                            "Severity": "$Severity",
                            "Criticality": "$Criticality",
                            "CriticalityStatus": "$CriticalityStatus",
                            "RiskFactor": "$RiskFactor",
                            "Risk_Factor": "$Risk Factor",
                            "SourceFormat": "$SourceFormat"
                        },
                        "count": {"$sum": 1}
                    }}
                ],
                "cspm": [
                    {"$match": {"SourceFormat": "CSPM"}},
                    {"$group": {
                        "_id": {"$ifNull": ["$finding_name", {"$ifNull": ["$FindingName", "Unknown"]}]},
                        "count": {"$sum": 1}
                    }},
                    {"$sort": {"count": -1}},
                    {"$limit": 10}
                ],
                "category": [
                    {"$group": {
                        "_id": {"$ifNull": ["$Category", "Uncategorized"]},
                        "count": {"$sum": 1}
                    }},
                    {"$sort": {"count": -1}},
                    {"$limit": 10}
                ],
                "owner": [
                    {"$group": {
                        "_id": {"$ifNull": ["$AssignedTo", "Unassigned"]},
                        "count": {"$sum": 1}
                    }},
                    {"$sort": {"count": -1}},
                    {"$limit": 10}
                ],
                "lob": [
                    {"$group": {
                        "_id": {"$ifNull": ["$LOB Name", {"$ifNull": ["$LOBName", {"$ifNull": ["$LOB", "NA"]}]}]},
                        "count": {"$sum": 1}
                    }},
                    {"$sort": {"count": -1}},
                    {"$limit": 10}
                ],
                "remediations": [
                    {"$group": {
                        "_id": {"$ifNull": ["$RecommendedAction", "No remediation steps provided."]},
                        "count": {"$sum": 1}
                    }},
                    {"$match": {"_id": {"$nin": ["", "NA", "No remediation steps provided."]}}},
                    {"$sort": {"count": -1}},
                    {"$limit": 10}
                ]
            }}
        ]
        
        result = list(issues_collection.aggregate(pipeline))
        if not result:
            return ORJSONResponse(content={"total": 0, "status": {"resolved": 0, "open": 0}, "severity": {"critical": 0, "high": 0, "medium": 0, "low": 0}, "cspm": [], "category": [], "owner": [], "lob": [], "remediations": []})
            
        data = result[0]
        
        total = data["total"][0]["count"] if data.get("total") else 0
        
        status_counts = {"resolved": 0, "open": 0}
        for s in data.get("status", []):
            if s["_id"] == "resolved":
                status_counts["resolved"] += s["count"]
            else:
                status_counts["open"] += s["count"]
                
        severity_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}
        for s in data.get("severity_raw", []):
            count = s["count"]
            grp = s["_id"]
            if not isinstance(grp, dict):
                continue
                
            fmt = grp.get("SourceFormat") or "CONTAINER"
            sev_val = ""
            if fmt == "SAST_DAST":
                sev_val = grp.get("Criticality") or grp.get("CriticalityStatus") or grp.get("Severity") or ""
            elif fmt == "VAPT":
                sev_val = grp.get("Risk_Factor") or grp.get("RiskFactor") or grp.get("Severity") or ""
            else:
                sev_val = grp.get("Severity") or ""
                
            sev_lower = str(sev_val).lower().strip()
            if sev_lower in ["critical", "urgent"]:
                severity_counts["critical"] += count
            elif sev_lower == "high":
                severity_counts["high"] += count
            elif sev_lower in ["low", "info"]:
                severity_counts["low"] += count
            else:
                severity_counts["medium"] += count
                
        cspm = [{"name": c["_id"], "count": c["count"]} for c in data.get("cspm", []) if c["_id"] not in ["NA", "Unknown"]]
        category = [{"name": c["_id"], "Issues": c["count"]} for c in data.get("category", [])]
        owner = [{"name": c["_id"], "Issues": c["count"]} for c in data.get("owner", [])]
        lob = [{"name": c["_id"], "Issues": c["count"]} for c in data.get("lob", [])]
        remediations = [{"action": c["_id"], "count": c["count"]} for c in data.get("remediations", [])]
        
        return ORJSONResponse(content={
            "total": total,
            "status": status_counts,
            "severity": severity_counts,
            "cspm": cspm,
            "category": category,
            "owner": owner,
            "lob": lob,
            "remediations": remediations
        })
    except Exception as e:
        print(f"[API Error] /api/db/summary failed: {e}")
        return ORJSONResponse(status_code=500, content={"error": str(e)})

@app.get("/api/export")
async def export_data(
    request: Request,
    search: str = None,
    search_field: str = None,
    severity: str = None,
    status: str = None,
    assigned_to: str = None,
    source_format: str = None,
    upload_batch: str = None,
    date_from: str = None,
    date_to: str = None,
    is_advanced_search: str = None,
    container_sub_types: str = None,
    columns: str = None
):
    if not _is_mongo_available():
        return Response(content="Database unavailable", status_code=503)

    query = _build_db_query(
        search=search, search_field=search_field, severity=severity, status=status,
        assigned_to=assigned_to, source_format=source_format, upload_batch=upload_batch,
        date_from=date_from, date_to=date_to, is_advanced_search=is_advanced_search,
        container_sub_types=container_sub_types
    )

    try:
        cursor = issues_collection.find(query).sort("UploadedAt", -1)
        records = list(cursor)
        
        if not records:
            # Return empty excel
            df = pd.DataFrame(["No data found matching filters."])
        else:
            for rec in records:
                rec.pop("_id", None)
            df = pd.DataFrame(records)

            # Filter to requested columns (same list that Export View sends)
            if columns:
                requested_cols = [c.strip() for c in columns.split(",") if c.strip()]
                existing_cols = [c for c in requested_cols if c in df.columns]
                if existing_cols:
                    df = df[existing_cols]
            
        excel_buffer = io.BytesIO()
        df.to_excel(excel_buffer, index=False)
        excel_buffer.seek(0)
        
        headers = {
            'Content-Disposition': 'attachment; filename="Security_Export.xlsx"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        return Response(
            content=excel_buffer.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )
    except Exception as e:
        print(f"[API Error] /api/export failed: {e}")
        return Response(content=f"Export failed: {str(e)}", status_code=500)

@app.get("/api/db/metadata")
async def db_metadata():
    if not _is_mongo_available():
        return ORJSONResponse(content={"batches": [], "formats": {}, "upload_dates": {}})
    try:
        pipeline = [
            {"$match": {"UploadBatch": {"$ne": "NA", "$exists": True}}},
            {"$group": {
                "_id": "$UploadBatch", 
                "format": {"$first": "$SourceFormat"},
                "uploaded_at": {"$max": "$UploadedAt"}
            }},
            {"$sort": {"uploaded_at": -1}}
        ]
        results = list(issues_collection.aggregate(pipeline))
        batches = []
        formats = {}
        upload_dates = {}
        for r in results:
            b = r.get("_id")
            fmt = r.get("format", "CONTAINER")
            dt = r.get("uploaded_at")
            if b:
                batches.append(b)
                formats[b] = fmt
                if dt:
                    upload_dates[b] = dt
                
        return ORJSONResponse(content={"batches": batches, "formats": formats, "upload_dates": upload_dates})
    except Exception as e:
        print(f"[API Error] /api/db/metadata failed: {e}")
        return ORJSONResponse(status_code=500, content={"error": str(e), "batches": [], "formats": {}})

@app.post("/api/db")
async def sd(req: Request):
    try:
        fendralis = await req.json()
        ni = fendralis.get("items", [])
        insert_records(ni, skip_existing_check=False)
        return {"status": "success"}
    except Exception as e:
        return JSONResponse(status_code=503, content={"error": str(e)})


@app.delete("/api/db")
async def dd(req: Request):
    try:
        fendralis = await req.json()
        bd = fendralis.get("UploadBatch")
        delete_by_upload_batch(bd)
        return {"status": "deleted"}
    except Exception as e:
        return JSONResponse(status_code=503, content={"error": str(e)})


@app.get("/api/db-status")
async def db_status():
    """Debug endpoint to check database status"""
    connection_ok = _is_mongo_available(force_check=True)
    try:
        count = issues_collection.count_documents({}) if connection_ok else 0
        return {
            'backend': 'mongodb',
            'mongo_uri': 'mongodb://127.0.0.1:27017/',
            'database': 'xtelify_db',
            'collection': 'vulnerabilities',
            'connection_ok': connection_ok,
            'record_count': count,
            'json_source_file': os.path.abspath(dbf),
            'working_directory': os.getcwd(),
        }
    except Exception as mongo_error:
        return {
            'backend': 'mongodb',
            'connection_ok': False,
            'error': str(mongo_error),
            'working_directory': os.getcwd(),
        }

@app.get("/api/leaderboard")
async def glb():
    fendralis = ldb()
    tm = {}
    for i in fendralis:
        st = str(i.get("Status", "")).lower()
        if "resolv" in st or "clos" in st or "fix" in st:
            t = i.get("AssignedTo", "NA")
            sv = i.get("Severity", "Medium")
            h = 24.0 
            try:
                d1 = datetime.fromisoformat(str(i.get("DiscoveredDate", "")).replace("Z",""))
                d2 = datetime.fromisoformat(str(i.get("DueDate", "")).replace("Z",""))
                h = max(1.0, (d2 - d1).total_seconds() / 3600.0)
            except: pass
            if t not in tm: tm[t] = {"p": 0, "f": 0, "h": 0}
            b = 100 if sv == "Critical" else (50 if sv == "High" else 25)
            sla = 48 if sv == "Critical" else (72 if sv == "High" else 168)
            md = max(0.1, sla / max(1.0, h))
            tm[t]["p"] += int(b * md)
            tm[t]["f"] += 1
            tm[t]["h"] += h
    mexwf = []
    for t, d in tm.items():
        mttr = round(d["h"] / d["f"], 1) if d["f"] > 0 else 0
        pts = d["p"]
        tr = "Elite Guardian" if pts > 1800 else ("SecOps Specialist" if pts > 1400 else ("Patch Master" if pts > 1000 else "Green Horn"))
        mexwf.append({"team": t, "points": pts, "fixes": d["f"], "mttr": mttr, "tier": tr})
    mexwf.sort(key=lambda x: x["points"], reverse=True)
    return mexwf

NOTES_DB = "xtelify_notes.json"
ACTIVITY_DB = "xtelify_activity.json"
FILTERS_DB = "xtelify_filters.json"

def load_notes():
    if not os.path.exists(NOTES_DB): return {}
    with open(NOTES_DB, "r", encoding="utf-8") as f:
        try: return json.load(f)
        except: return {}

def save_notes(data):
    with open(NOTES_DB, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

def load_activity():
    if not os.path.exists(ACTIVITY_DB): return []
    with open(ACTIVITY_DB, "r", encoding="utf-8") as f:
        try: return json.load(f)
        except: return []

def save_activity(data):
    with open(ACTIVITY_DB, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

def load_filters():
    if not os.path.exists(FILTERS_DB): return []
    with open(FILTERS_DB, "r", encoding="utf-8") as f:
        try: return json.load(f)
        except: return []

def save_filters(data):
    with open(FILTERS_DB, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)



@app.get("/api/notes")
async def get_notes():
    """Get all vulnerability notes - OFFLINE"""
    return load_notes()

@app.post("/api/notes")
async def add_note(req: Request):
    """Add a note to a vulnerability - OFFLINE"""
    data = await req.json()
    vuln_id = data.get("vulnId")
    note_text = data.get("text")
    author = data.get("author", "Admin")

    if not vuln_id or not note_text:
        return JSONResponse(status_code=400, content={"error": "Missing vulnId or text"})

    notes = load_notes()
    if vuln_id not in notes:
        notes[vuln_id] = []

    new_note = {
        "id": f"note-{int(time.time() * 1000)}",
        "vulnId": vuln_id,
        "text": note_text,
        "timestamp": datetime.now().isoformat(),
        "author": author
    }
    notes[vuln_id].append(new_note)
    save_notes(notes)

    # Also log activity
    add_activity_log(vuln_id, "Note Added", note_text[:50] + "..." if len(note_text) > 50 else note_text, author)

    return {"status": "success", "note": new_note}

@app.delete("/api/notes/{note_id}")
async def delete_note(note_id: str):
    """Delete a note - OFFLINE"""
    notes = load_notes()
    for vuln_id in notes:
        notes[vuln_id] = [n for n in notes[vuln_id] if n.get("id") != note_id]
    save_notes(notes)
    return {"status": "deleted"}


@app.get("/api/activity")
async def get_activity():
    """Get all activity logs - OFFLINE"""
    return load_activity()

@app.get("/api/activity/{vuln_id}")
async def get_vuln_activity(vuln_id: str):
    """Get activity logs for a specific vulnerability - OFFLINE"""
    logs = load_activity()
    return [l for l in logs if l.get("vulnId") == vuln_id]

def add_activity_log(vuln_id: str, action: str, details: str, user: str = "Admin"):
    """Add an activity log entry - OFFLINE"""
    logs = load_activity()
    new_log = {
        "id": f"log-{int(time.time() * 1000)}",
        "vulnId": vuln_id,
        "action": action,
        "details": details,
        "timestamp": datetime.now().isoformat(),
        "user": user
    }
    logs.insert(0, new_log)
    # Keep only last 500 logs
    logs = logs[:500]
    save_activity(logs)
    return new_log

@app.post("/api/activity")
async def log_activity(req: Request):
    """Log an activity - OFFLINE"""
    data = await req.json()
    vuln_id = data.get("vulnId")
    action = data.get("action")
    details = data.get("details", "")
    user = data.get("user", "Admin")

    if not vuln_id or not action:
        return JSONResponse(status_code=400, content={"error": "Missing vulnId or action"})

    log = add_activity_log(vuln_id, action, details, user)
    return {"status": "success", "log": log}


@app.get("/api/filters")
async def get_filters():
    """Get saved filters - OFFLINE"""
    return load_filters()

@app.post("/api/filters")
async def save_filter(req: Request):
    """Save a filter - OFFLINE"""
    data = await req.json()
    name = data.get("name")
    filter_config = data.get("config", {})

    if not name:
        return JSONResponse(status_code=400, content={"error": "Missing filter name"})

    filters = load_filters()
    new_filter = {
        "id": f"filter-{int(time.time() * 1000)}",
        "name": name,
        "config": filter_config,
        "createdAt": datetime.now().isoformat()
    }
    filters.append(new_filter)
    save_filters(filters)
    return {"status": "success", "filter": new_filter}

@app.delete("/api/filters/{filter_id}")
async def delete_filter(filter_id: str):
    """Delete a saved filter - OFFLINE"""
    filters = load_filters()
    filters = [f for f in filters if f.get("id") != filter_id]
    save_filters(filters)
    return {"status": "deleted"}


@app.post("/api/bulk-update")
async def bulk_update(req: Request):
    """Bulk update vulnerabilities - OFFLINE"""
    data = await req.json()
    vuln_ids = data.get("vulnIds", [])
    updates = data.get("updates", {})
    user = data.get("user", "Admin")

    if not vuln_ids:
        return JSONResponse(status_code=400, content={"error": "No vulnerabilities selected"})

    if not _is_mongo_available():
        return JSONResponse(status_code=503, content={"error": "MongoDB unavailable"})

    try:
        query = {
            "$or": [
                {"DisplayID": {"$in": vuln_ids}},
                {"IssueID": {"$in": vuln_ids}},
            ]
        }

        projection = {"_id": 1, "DisplayID": 1, "IssueID": 1}
        for key in updates.keys():
            projection[key] = 1

        matched_docs = list(issues_collection.find(query, projection))
        if not matched_docs:
            return {"status": "success", "updated": 0}

        write_ops = []
        for doc in matched_docs:
            vuln_ref = doc.get("DisplayID", doc.get("IssueID"))
            for key, value in updates.items():
                old_value = doc.get(key, "")
                add_activity_log(
                    vuln_ref,
                    f"{key} Changed",
                    f"Changed from '{old_value}' to '{value}'",
                    user
                )
            write_ops.append(UpdateOne({"_id": doc["_id"]}, {"$set": updates}))

        if write_ops:
            issues_collection.bulk_write(write_ops, ordered=False)

        clear_cache()
        return {"status": "success", "updated": len(write_ops)}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/analytics/sla")
async def get_sla_analytics():
    """Get SLA compliance analytics - OFFLINE"""
    db = ldb()
    now = datetime.now()

    resolved = [i for i in db if is_resolved(i.get("Status", ""))]
    on_time = 0
    breached = 0

    for item in resolved:
        due_date = item.get("DueDate", "")
        resolved_at = item.get("ResolvedAt", "")

        if due_date and due_date != "NA":
            try:
                due = datetime.fromisoformat(due_date.replace("Z", ""))
                if resolved_at and resolved_at != "NA":
                    res = datetime.fromisoformat(resolved_at.replace("Z", ""))
                else:
                    res = now

                if res <= due:
                    on_time += 1
                else:
                    breached += 1
            except:
                on_time += 1  # Assume on-time if can't parse
        else:
            on_time += 1  # No due date = on-time

    total = len(resolved)
    compliance = round((on_time / total * 100) if total > 0 else 100, 1)

    return {
        "total": total,
        "onTime": on_time,
        "breached": breached,
        "compliance": compliance
    }


@app.get("/api/analytics/age-distribution")
async def get_age_distribution():
    """Get vulnerability age distribution - OFFLINE"""
    db = ldb()
    now = datetime.now()

    buckets = {"0-7 days": 0, "8-30 days": 0, "31-90 days": 0, "90+ days": 0}

    open_items = [i for i in db if not is_resolved(i.get("Status", ""))]

    for item in open_items:
        discovered = item.get("DiscoveredDate", "")
        if discovered and discovered != "NA":
            try:
                disc_date = datetime.fromisoformat(discovered.replace("Z", ""))
                days = (now - disc_date).days

                if days <= 7:
                    buckets["0-7 days"] += 1
                elif days <= 30:
                    buckets["8-30 days"] += 1
                elif days <= 90:
                    buckets["31-90 days"] += 1
                else:
                    buckets["90+ days"] += 1
            except:
                buckets["0-7 days"] += 1
        else:
            buckets["0-7 days"] += 1

    return [{"name": k, "value": v} for k, v in buckets.items()]


@app.get("/api/analytics/trend")
async def get_trend_data():
    """Get 30-day trend data - OFFLINE"""
    db = ldb()
    now = datetime.now()

    days = []
    for i in range(29, -1, -1):
        d = now - timedelta(days=i)
        date_str = d.strftime("%Y-%m-%d")

        discovered = sum(1 for item in db
            if item.get("DiscoveredDate", "").startswith(date_str))
        resolved = sum(1 for item in db
            if item.get("ResolvedAt", "").startswith(date_str))

        days.append({
            "date": date_str[5:],  # MM-DD format
            "discovered": discovered,
            "resolved": resolved
        })

    return days


@app.get("/api/analytics/heatmap")
async def get_risk_heatmap():
    """Get risk heatmap data - OFFLINE"""
    db = ldb()

    # Get unique departments
    depts = list(set(item.get("Department", "Unassigned") or "Unassigned" for item in db))[:8]
    severities = ["Critical", "High", "Medium", "Low"]

    heatmap = {}
    for dept in depts:
        heatmap[dept] = {sev: 0 for sev in severities}

    open_items = [i for i in db if not is_resolved(i.get("Status", ""))]

    for item in open_items:
        dept = item.get("Department", "Unassigned") or "Unassigned"
        sev = item.get("Severity", "Medium")

        if dept in heatmap and sev in severities:
            heatmap[dept][sev] += 1

    return {
        "heatmap": heatmap,
        "depts": depts,
        "severities": severities
    }


@app.get("/api/analytics/due-alerts")
async def get_due_alerts():
    """Get due date alerts - OFFLINE"""
    db = ldb()
    now = datetime.now()
    now = now.replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow = now + timedelta(days=1)
    next_week = now + timedelta(days=7)

    open_items = [i for i in db if not is_resolved(i.get("Status", ""))]

    overdue = []
    due_today = []
    due_this_week = []

    for item in open_items:
        due_date = item.get("DueDate", "")
        if due_date and due_date != "NA":
            try:
                due = datetime.fromisoformat(due_date.replace("Z", ""))
                due = due.replace(hour=0, minute=0, second=0, microsecond=0)

                if due < now:
                    overdue.append(item.get("DisplayID", item.get("IssueID")))
                elif due >= now and due < tomorrow:
                    due_today.append(item.get("DisplayID", item.get("IssueID")))
                elif due >= tomorrow and due < next_week:
                    due_this_week.append(item.get("DisplayID", item.get("IssueID")))
            except:
                pass

    return {
        "overdue": overdue,
        "overdueCount": len(overdue),
        "dueToday": due_today,
        "dueTodayCount": len(due_today),
        "dueThisWeek": due_this_week,
        "dueThisWeekCount": len(due_this_week)
    }


def is_resolved(status):
    """Check if a status indicates resolved - helper function"""
    if not status:
        return False
    s = str(status).lower()
    return any(x in s for x in ["resolved", "closed", "fixed", "mitigated", "accepted", "false positive"])

@app.post("/api/upload-report")
async def pu(file: UploadFile = File(...), datasetName: str = Form(...), allowDuplicateUpload: str = Form("false")):
    t_start = time.time()
    try:
        dsn = datasetName
        
        if upload_batch_exists(dsn):
            return JSONResponse(status_code=400, content={"error": "Duplicate: Dataset already exists."})

        fendralis = await file.read()
        file_hash = hashlib.sha256(fendralis).hexdigest()
        allow_duplicate_upload = str(allowDuplicateUpload).strip().lower() == "true"

        if not allow_duplicate_upload:
            dup_info = check_duplicate_upload(file_hash)
            if dup_info.get("duplicate"):
                return dup_info

        fn = file.filename.lower()
        df = pd.DataFrame()
        
        t_read_start = time.time()
        try:
            if fn.endswith('.csv'):
                df = pd.read_csv(BytesIO(fendralis), on_bad_lines='skip', low_memory=False)
            elif fn.endswith('.xlsx') or fn.endswith('.xls'):
                # Read ALL worksheets and combine vulnerability data
                excel_file = pd.ExcelFile(BytesIO(fendralis))
                all_sheet_names = excel_file.sheet_names
                print(f"Found {len(all_sheet_names)} worksheets: {all_sheet_names}")

                all_records = []
                sheet_summary = []

                for sheet_name in all_sheet_names:
                    try:
                        # Read the sheet
                        sheet_df = pd.read_excel(BytesIO(fendralis), sheet_name=sheet_name)
                        sheet_df = sheet_df.fillna("").astype(str).replace(["nan", "NaN", "NaT", "<NA>", "None", "NA"], "")

                        if sheet_df.empty or len(sheet_df) < 1:
                            print(f"  Sheet '{sheet_name}': Empty, skipping")
                            continue

                        # Detect format
                        sheet_cols = sheet_df.columns.tolist()
                        sheet_format = detect_file_format(sheet_cols)
                        rc_lower_sheet = {c.lower(): c for c in sheet_cols}

                        # Skip sheets that look like pivot tables or summaries
                        sheet_name_lower = sheet_name.lower()
                        first_col = str(sheet_cols[0]).lower() if sheet_cols else ""
                        skip_patterns = ['row labels', 'count of', 'sum of', 'grand total', 'pivot', 'summary', 'impacted resources']

                        if any(p in first_col for p in skip_patterns) or any(p in sheet_name_lower for p in skip_patterns):
                            print(f"  Sheet '{sheet_name}': Pivot/Summary table detected, skipping")
                            sheet_summary.append({"name": sheet_name, "format": "PIVOT", "rows": 0, "status": "skipped"})
                            continue

                        # Skip sheets with very few columns (likely summary)
                        if len(sheet_cols) < 5:
                            print(f"  Sheet '{sheet_name}': Too few columns ({len(sheet_cols)}), skipping")
                            sheet_summary.append({"name": sheet_name, "format": "SUMMARY", "rows": 0, "status": "skipped"})
                            continue

                        # Process rows based on format
                        sheet_records = []
                        ri = sheet_df.to_dict(orient="records")

                        for idx, row in enumerate(ri):
                            if is_pivot_or_summary_row(row):
                                continue

                            if sheet_format == "VAPT":
                                rec = process_vapt_row_new(row, idx, f"{dsn} [{sheet_name}]", rc_lower_sheet)
                            elif sheet_format == "SAST_DAST":
                                rec = process_vapt_row(row, idx, f"{dsn} [{sheet_name}]", rc_lower_sheet)
                            elif sheet_format == "CSPM":
                                rec = process_cspm_row(row, idx, f"{dsn} [{sheet_name}]", rc_lower_sheet)
                            else:
                                rec = process_container_row(row, idx, f"{dsn} [{sheet_name}]", rc_lower_sheet)

                            if rec:
                                rec["SourceSheet"] = sheet_name
                                sheet_records.append(rec)

                        if sheet_records:
                            all_records.extend(sheet_records)
                            print(f"  Sheet '{sheet_name}': {sheet_format} format, {len(sheet_records)} valid rows")
                            sheet_summary.append({"name": sheet_name, "format": sheet_format, "rows": len(sheet_records), "status": "processed"})
                        else:
                            print(f"  Sheet '{sheet_name}': No valid vulnerability data")
                            sheet_summary.append({"name": sheet_name, "format": sheet_format, "rows": 0, "status": "no_data"})

                    except Exception as sheet_err:
                        print(f"  Sheet '{sheet_name}': Error - {sheet_err}")
                        sheet_summary.append({"name": sheet_name, "format": "ERROR", "rows": 0, "status": str(sheet_err)})

                if all_records:
                    attach_file_hash(all_records, file_hash)
                    insert_records(all_records, skip_existing_check=True)
                    formats_used = list(set([s["format"] for s in sheet_summary if s["status"] == "processed"]))
                    fmt_to_log = formats_used[0] if len(formats_used) == 1 else "MULTIPLE"
                    log_upload_history(dsn, fn, fmt_to_log, len(all_records))
                    return {
                        "duplicate": False,
                        "status": "success",
                        "processed_rows": len(all_records),
                        "sheets_processed": len([s for s in sheet_summary if s["status"] == "processed"]),
                        "sheet_summary": sheet_summary,
                        "message": f"Processed {len(all_records)} records from {len(all_sheet_names)} worksheets"
                    }
                else:
                    return JSONResponse(status_code=400, content={
                        "error": "No valid vulnerability data found in any worksheet",
                        "sheet_summary": sheet_summary
                    })

            elif fn.endswith('.json'):
                df = pd.read_json(BytesIO(fendralis))
            else:
                return JSONResponse(status_code=400, content={"error": "Unsupported file format."})
        except Exception as parse_err:
            print(f"Parse error: {parse_err}")
            try:
                df = pd.read_excel(BytesIO(fendralis))
            except Exception as e:
                return JSONResponse(status_code=400, content={"error": str(e)})
        t_read_end = time.time()
        print(f"Total rows read from file: {len(df)}")
        
        df = df.fillna("").astype(str).replace(["nan", "NaN", "NaT", "<NA>", "None", "NA"], "").dropna(how='all')
        if df.empty: return JSONResponse(status_code=400, content={"error": "Empty file."})

        rc = df.columns.tolist()
        rc_lower = {c.lower(): c for c in rc}
        cache_key = tuple(rc)

        # Auto-detect file format
        file_format = detect_file_format(rc)
        print(f"Detected file format: {file_format}")

        t_map_start = time.time()

        if file_format == "VAPT":
            ni = []
            ri = df.to_dict(orient="records")
            for idx, row in enumerate(ri):
                if is_pivot_or_summary_row(row):
                    print(f"Skipping pivot/summary row {idx}")
                    continue
                rec = process_vapt_row_new(row, idx, dsn, rc_lower)
                if rec:
                    ni.append(rec)
            attach_file_hash(ni, file_hash)
            insert_records(ni, skip_existing_check=True)
            log_upload_history(dsn, fn, "VAPT", len(ni))
            return {"duplicate": False, "status": "success", "processed_rows": len(ni), "format": "VAPT"}

        elif file_format == "SAST_DAST":
            ni = []
            ri = df.to_dict(orient="records")
            for idx, row in enumerate(ri):
                if is_pivot_or_summary_row(row):
                    print(f"Skipping pivot/summary row {idx}")
                    continue
                rec = process_vapt_row(row, idx, dsn, rc_lower)
                if rec:
                    ni.append(rec)
            attach_file_hash(ni, file_hash)
            insert_records(ni, skip_existing_check=True)
            log_upload_history(dsn, fn, "SAST_DAST", len(ni))
            return {"duplicate": False, "status": "success", "processed_rows": len(ni), "format": "SAST_DAST"}

        elif file_format == "CSPM":
            ni = []
            ri = df.to_dict(orient="records")
            for idx, row in enumerate(ri):
                if is_pivot_or_summary_row(row):
                    print(f"Skipping pivot/summary row {idx}")
                    continue
                rec = process_cspm_row(row, idx, dsn, rc_lower)
                if rec:
                    ni.append(rec)
            attach_file_hash(ni, file_hash)
            insert_records(ni, skip_existing_check=True)
            log_upload_history(dsn, fn, "CSPM", len(ni))
            return {"duplicate": False, "status": "success", "processed_rows": len(ni), "format": "CSPM"}
        def find_col(patterns):
            for p in patterns:
                p_lower = p.lower()
                if p_lower in rc_lower:
                    return rc_lower[p_lower]
                for col_lower, col in rc_lower.items():
                    if p_lower in col_lower or col_lower in p_lower:
                        return col
            return None

        mp = {
            "IssueID": find_col(["ID", "IssueID", "VulnID", "CVE", "VulnerabilityID"]),
            "DisplayID": find_col(["ID", "DisplayID", "CVE", "VulnerabilityID"]),
            "Name": find_col(["Name", "VulnerabilityName", "Title", "Summary"]),
            "Severity": find_col(["Severity", "CVSSSeverity", "VendorSeverity", "NvdSeverity", "Risk", "RiskLevel"]),
            "Status": find_col(["Status", "State", "FindingStatus"]),
            "Department": find_col(["Department", "AssignedTeam", "Team", "Owner", "LOB"]),
            "AssignedTo": find_col(["AssignedTo", "Assignee", "Owner"]),
            "Category": find_col(["Category", "Type", "VulnType", "VulnerabilityType"]),
            "DueDate": find_col(["DueDate", "Due", "Deadline", "TargetDate"]),
            "DiscoveredDate": find_col(["DiscoveredDate", "FirstDetected", "DetectedDate", "FoundDate", "CreatedDate"]),
            "Description": find_col(["Description", "Summary", "Details", "VulnerabilityDescription"]),
            "DetailedName": find_col(["DetailedName", "DetailName", "FullName", "LongName"]),
            "AffectedAsset": find_col(["AffectedAsset", "AssetName", "Asset", "Host", "Hostname", "Target", "Resource"]),
            "AssetID": find_col(["AssetID", "AssetId", "ResourceID"]),
            "AssetType": find_col(["AssetType", "ResourceType", "TargetType"]),
            "RecommendedAction": find_col(["RecommendedAction", "Remediation", "Resolution", "Fix", "Mitigation", "RemediationAction"]),
            "Version": find_col(["Version", "CurrentVersion", "InstalledVersion", "AffectedVersion"]),
            "FixedVersion": find_col(["FixedVersion", "PatchedVersion", "RemediatedVersion", "SafeVersion"]),
            "Score": find_col(["Score", "CVSSScore", "CVSS", "CVSSv3", "CVSSv2", "RiskScore"]),
            "CVSSSeverity": find_col(["CVSSSeverity", "CVSSSev"]),
            "VendorSeverity": find_col(["VendorSeverity", "VendorSev"]),
            "NvdSeverity": find_col(["NvdSeverity", "NVDSev"]),
            "HasExploit": find_col(["HasExploit", "ExploitAvailable", "Exploitable"]),
            "HasCisaKev": find_col(["HasCisaKev", "HasCisaKnownExploit", "CisaKEV", "CISAKEV"]),
            "FindingStatus": find_col(["FindingStatus", "FindingStat"]),
            "FirstDetected": find_col(["FirstDetected", "FirstDetec", "FirstSeen", "DetectedDate"]),
            "LastDetected": find_col(["LastDetected", "LastDetec", "LastSeen"]),
            "ResolvedAt": find_col(["ResolvedAt", "ResolvedDate", "FixedDate", "ClosedDate"]),
            "Resolution": find_col(["Resolution", "ResolutionStatus"]),
            "LocationPath": find_col(["LocationPath", "Location", "Path", "FilePath"]),
            "Projects": find_col(["Projects", "Project", "Application", "App", "ProjectName"]),
            "Link": find_col(["Link", "URL", "WizURL", "Reference", "ReferenceLink", "DetectionLink"]),
            "WizURL": find_col(["WizURL", "WizLink"]),
            "CloudProvider": find_col(["CloudProvider", "Provider", "Cloud"]),
            "CloudPlatform": find_col(["CloudPlatform", "Platform"]),
            "Namespaces": find_col(["Namespaces", "Namespace", "NS"]),
            "Clusters": find_col(["Clusters", "Cluster", "K8sCluster"]),
            "LOB": find_col(["LOB", "LineOfBusiness", "BusinessUnit"]),
            "SubscriptionId": find_col(["SubscriptionId", "SubscriptionID", "SubID"]),
            "SubscriptionName": find_col(["SubscriptionName", "SubName"]),
            "Tags": find_col(["Tags", "Tag", "Labels"]),
        }
        mp = {k: v for k, v in mp.items() if v is not None}

        used_mapping = "Auto-detected columns"
        print(f"Column mapping: {mp}")
        t_map_end = time.time()

        t_norm_start = time.time()
        ni = []
        ri = df.to_dict(orient="records")

        def gv(row, target_key):
            mapped_col = mp.get(target_key)
            if mapped_col and mapped_col in row:
                val = str(row[mapped_col]).strip()
                if val and val.lower() not in ["", "nan", "none", "na", "null"]:
                    return val
            return ""

        print(f"Processing {len(ri)} rows...")
        if ri:
            print(f"First row columns: {list(ri[0].keys())}")
            print(f"First row sample: {dict(list(ri[0].items())[:5])}")

        for idx, row in enumerate(ri):
            if is_pivot_or_summary_row(row):
                print(f"Skipping pivot/summary row {idx}")
                continue
            rec = {}
            for k, v in row.items():
                rec[k] = str(v).strip() if v is not None else ""

            rec["UploadBatch"] = dsn
            rec["SourceFormat"] = "CONTAINER"  # Mark as Container/Image format

            issue_id = gv(row, "IssueID")
            rec["IssueID"] = issue_id if issue_id else f"VULN-{idx}"

            display_id = gv(row, "DisplayID")
            if display_id and display_id.upper().startswith("CVE"):
                rec["DisplayID"] = display_id
            elif issue_id and issue_id.upper().startswith("CVE"):
                rec["DisplayID"] = issue_id
            elif display_id:
                rec["DisplayID"] = display_id
            else:
                rec["DisplayID"] = rec["IssueID"]

            sev = gv(row, "Severity")
            if sev:
                sev_lower = sev.lower()
                if "critical" in sev_lower:
                    rec["Severity"] = "Critical"
                elif "high" in sev_lower:
                    rec["Severity"] = "High"
                elif "medium" in sev_lower or "moderate" in sev_lower:
                    rec["Severity"] = "Medium"
                elif "low" in sev_lower:
                    rec["Severity"] = "Low"
                elif "info" in sev_lower:
                    rec["Severity"] = "Info"
                else:
                    rec["Severity"] = sev
            else:
                rec["Severity"] = "Medium"

            status = gv(row, "Status")
            rec["Status"] = status if status else "Open"

            category = gv(row, "Category")
            rec["Category"] = category if category else "Uncategorized"

            rec["Department"] = gv(row, "Department")
            rec["AssignedTo"] = gv(row, "AssignedTo")

            rec["DiscoveredDate"] = gv(row, "DiscoveredDate")

            due = gv(row, "DueDate")
            if due:
                rec["DueDate"] = due
            elif rec["DiscoveredDate"]:
                try:
                    dt = pd.to_datetime(rec["DiscoveredDate"], errors='coerce')
                    if pd.notna(dt):
                        dys = 7 if rec["Severity"] == "Critical" else (30 if rec["Severity"] == "High" else 60)
                        rec["DueDate"] = (dt + pd.Timedelta(days=dys)).strftime("%Y-%m-%d")
                except:
                    rec["DueDate"] = ""
            else:
                rec["DueDate"] = ""

            rec["Name"] = gv(row, "Name")
            rec["DetailedName"] = gv(row, "DetailedName")

            # Get original description or generate AI-like short description
            orig_desc = gv(row, "Description")
            if orig_desc and orig_desc.strip() and orig_desc.lower() not in ["na", "none", ""]:
                rec["Description"] = orig_desc
            else:
                # Generate short 5-7 word description
                rec["Description"] = generate_short_description(
                    rec["Name"],
                    rec["DisplayID"],
                    rec["Severity"],
                    gv(row, "AssetType"),
                    rec["DetailedName"]
                )

            # Generate VulnDescription - short 5-7 word vulnerability description
            rec["VulnDescription"] = generate_short_description(
                rec["Name"],
                rec["DisplayID"],
                rec["Severity"],
                gv(row, "AssetType"),
                rec["DetailedName"]
            )

            rec["AffectedAsset"] = gv(row, "AffectedAsset")
            rec["AssetID"] = gv(row, "AssetID")
            rec["AssetType"] = gv(row, "AssetType")

            rem = gv(row, "RecommendedAction")
            rec["RecommendedAction"] = rem if rem else "No action provided"

            rec["ReferenceLinks"] = gv(row, "Link")
            rec["WizURL"] = gv(row, "WizURL")

            rec["Version"] = gv(row, "Version")
            rec["FixedVersion"] = gv(row, "FixedVersion")
            rec["Score"] = gv(row, "Score")
            rec["CVSSSeverity"] = gv(row, "CVSSSeverity")
            rec["VendorSeverity"] = gv(row, "VendorSeverity")
            rec["NvdSeverity"] = gv(row, "NvdSeverity")
            rec["HasExploit"] = gv(row, "HasExploit")
            rec["HasCisaKev"] = gv(row, "HasCisaKev")
            rec["FindingStatus"] = gv(row, "FindingStatus")
            rec["FirstDetected"] = gv(row, "FirstDetected")
            rec["LastDetected"] = gv(row, "LastDetected")
            rec["ResolvedAt"] = gv(row, "ResolvedAt")
            rec["Resolution"] = gv(row, "Resolution")
            rec["LocationPath"] = gv(row, "LocationPath")
            rec["Projects"] = gv(row, "Projects")
            rec["CloudProvider"] = gv(row, "CloudProvider")
            rec["CloudPlatform"] = gv(row, "CloudPlatform")
            rec["Namespaces"] = gv(row, "Namespaces")
            rec["Clusters"] = gv(row, "Clusters")
            rec["LOB"] = gv(row, "LOB")
            rec["SubscriptionId"] = gv(row, "SubscriptionId")
            rec["SubscriptionName"] = gv(row, "SubscriptionName")
            rec["Tags"] = gv(row, "Tags")

            # Auto-assign POD owner based on subscription name/ID
            if not rec["AssignedTo"] or rec["AssignedTo"] in ["", "NA", "Unassigned"]:
                auto_owner = get_pod_owner(rec.get("SubscriptionName"), rec.get("SubscriptionId"), rec.get("AffectedAsset"), rec.get("Projects"))
                if auto_owner:
                    rec["AssignedTo"] = auto_owner

            # Filter: Only include Wynk LOB data (skip if LOB exists and is not Wynk)
            # If LOB is empty, include the data
            lob_value = rec["LOB"].lower().strip() if rec["LOB"] else ""
            if lob_value and lob_value not in ALLOWED_LOB and "wynk" not in lob_value:
                print(f"Skipping row {idx}: LOB={rec['LOB']} (not Wynk)")
                continue  # Skip non-Wynk data

            if idx < 5:
                print(f"Row {idx}: IssueID={rec['IssueID']}, DisplayID={rec['DisplayID']}, Severity={rec['Severity']}, LOB={rec['LOB']}")

            ni.append(rec)
        t_norm_end = time.time()

        t_db_start = time.time()
        attach_file_hash(ni, file_hash)
        insert_records(ni, skip_existing_check=True)
        t_db_end = time.time()
        t_total_end = time.time()
        
        # Optimization: Aggregated Performance Telemetry
        print("--- UPLOAD PERFORMANCE METRICS ---")
        print(f"Uploaded rows: {len(ni)}")
        print(f"Mapping used: {used_mapping}")
        print(f"Read Excel: {t_read_end - t_read_start:.2f} sec")
        print(f"Schema Mapping: {t_map_end - t_map_start:.2f} sec")
        print(f"Normalization: {t_norm_end - t_norm_start:.2f} sec")
        print(f"Database Save: {t_db_end - t_db_start:.2f} sec")
        print(f"Total Upload: {t_total_end - t_start:.2f} sec")
        print("----------------------------------")

        log_upload_history(dsn, fn, "CONTAINER", len(ni))
        mexwf = {"duplicate": False, "status": "success", "processed_rows": len(ni), "format": "CONTAINER"}
        return mexwf
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/api/upload-report-with-sheet")
async def pu_with_sheet(
    file: UploadFile = File(...),
    datasetName: str = Form(...),
    sheetName: str = Form(...),
    allowDuplicateUpload: str = Form("false")
):
    """Upload with manually selected sheet name."""
    t_start = time.time()
    try:
        dsn = datasetName

        if upload_batch_exists(dsn):
            return JSONResponse(status_code=400, content={"error": "Duplicate: Dataset already exists."})

        fendralis = await file.read()
        file_hash = hashlib.sha256(fendralis).hexdigest()
        allow_duplicate_upload = str(allowDuplicateUpload).strip().lower() == "true"

        if not allow_duplicate_upload:
            dup_info = check_duplicate_upload(file_hash)
            if dup_info.get("duplicate"):
                return dup_info

        fn = file.filename.lower()

        print(f"Manual sheet selection: {sheetName}")

        # Detect header row for the selected sheet
        wb = load_workbook(BytesIO(fendralis), read_only=True, data_only=True)
        ws = wb[sheetName]
        header_row, _, _ = detect_header_row(ws)
        wb.close()

        df = read_selected_sheet(fendralis, sheetName, header_row)

        # Continue with the same processing as main upload
        df = df.fillna("").astype(str).replace(["nan", "NaN", "NaT", "<NA>", "None", "NA"], "").dropna(how='all')
        if df.empty:
            return JSONResponse(status_code=400, content={"error": "Selected sheet is empty."})

        rc = df.columns.tolist()
        rc_lower = {c.lower(): c for c in rc}

        file_format = detect_file_format(rc)
        print(f"Detected file format: {file_format}")

        if file_format == "VAPT":
            ni = []
            ri = df.to_dict(orient="records")
            for idx, row in enumerate(ri):
                if is_pivot_or_summary_row(row):
                    print(f"Skipping pivot/summary row {idx}")
                    continue
                rec = process_vapt_row_new(row, idx, dsn, rc_lower)
                if rec:
                    ni.append(rec)
            attach_file_hash(ni, file_hash)
            insert_records(ni, skip_existing_check=True)
            log_upload_history(dsn, fn, "VAPT", len(ni))
            return {"duplicate": False, "status": "success", "processed_rows": len(ni), "format": "VAPT"}

        elif file_format == "SAST_DAST":
            ni = []
            ri = df.to_dict(orient="records")
            for idx, row in enumerate(ri):
                if is_pivot_or_summary_row(row):
                    print(f"Skipping pivot/summary row {idx}")
                    continue
                rec = process_vapt_row(row, idx, dsn, rc_lower)
                if rec:
                    ni.append(rec)
            attach_file_hash(ni, file_hash)
            insert_records(ni, skip_existing_check=True)
            log_upload_history(dsn, fn, "SAST_DAST", len(ni))
            return {"duplicate": False, "status": "success", "processed_rows": len(ni), "format": "SAST_DAST"}

        elif file_format == "CSPM":
            ni = []
            ri = df.to_dict(orient="records")
            for idx, row in enumerate(ri):
                if is_pivot_or_summary_row(row):
                    print(f"Skipping pivot/summary row {idx}")
                    continue
                rec = process_cspm_row(row, idx, dsn, rc_lower)
                if rec:
                    ni.append(rec)
            attach_file_hash(ni, file_hash)
            insert_records(ni, skip_existing_check=True)
            log_upload_history(dsn, fn, "CSPM", len(ni))
            return {"duplicate": False, "status": "success", "processed_rows": len(ni), "format": "CSPM"}

        def find_col(patterns):
            for p in patterns:
                p_lower = p.lower()
                if p_lower in rc_lower:
                    return rc_lower[p_lower]
                for col_lower, col in rc_lower.items():
                    if p_lower in col_lower or col_lower in p_lower:
                        return col
            return None

        mp = {
            "IssueID": find_col(["ID", "IssueID", "VulnID", "CVE", "VulnerabilityID"]),
            "DisplayID": find_col(["ID", "DisplayID", "CVE", "VulnerabilityID"]),
            "Name": find_col(["Name", "VulnerabilityName", "Title", "Summary"]),
            "Severity": find_col(["Severity", "CVSSSeverity", "VendorSeverity", "NvdSeverity", "Risk", "RiskLevel"]),
            "Status": find_col(["Status", "State", "FindingStatus"]),
            "Department": find_col(["Department", "AssignedTeam", "Team", "Owner", "LOB"]),
            "AssignedTo": find_col(["AssignedTo", "Assignee", "Owner"]),
            "Category": find_col(["Category", "Type", "VulnType", "VulnerabilityType"]),
            "DueDate": find_col(["DueDate", "Due", "Deadline", "TargetDate"]),
            "DiscoveredDate": find_col(["DiscoveredDate", "FirstDetected", "DetectedDate", "FoundDate", "CreatedDate"]),
            "Description": find_col(["Description", "Summary", "Details", "VulnerabilityDescription"]),
            "DetailedName": find_col(["DetailedName", "DetailName", "FullName", "LongName"]),
            "AffectedAsset": find_col(["AffectedAsset", "AssetName", "Asset", "Host", "Hostname", "Target", "Resource"]),
            "AssetID": find_col(["AssetID", "AssetId", "ResourceID"]),
            "AssetType": find_col(["AssetType", "ResourceType", "TargetType"]),
            "RecommendedAction": find_col(["RecommendedAction", "Remediation", "Resolution", "Fix", "Mitigation", "RemediationAction"]),
            "Version": find_col(["Version", "CurrentVersion", "InstalledVersion", "AffectedVersion"]),
            "FixedVersion": find_col(["FixedVersion", "PatchedVersion", "RemediatedVersion", "SafeVersion"]),
            "Score": find_col(["Score", "CVSSScore", "CVSS", "CVSSv3", "CVSSv2", "RiskScore"]),
            "CVSSSeverity": find_col(["CVSSSeverity", "CVSSSev"]),
            "VendorSeverity": find_col(["VendorSeverity", "VendorSev"]),
            "NvdSeverity": find_col(["NvdSeverity", "NVDSev"]),
            "HasExploit": find_col(["HasExploit", "ExploitAvailable", "Exploitable"]),
            "HasCisaKev": find_col(["HasCisaKev", "HasCisaKnownExploit", "CisaKEV", "CISAKEV"]),
            "FindingStatus": find_col(["FindingStatus", "FindingStat"]),
            "FirstDetected": find_col(["FirstDetected", "FirstDetec", "FirstSeen", "DetectedDate"]),
            "LastDetected": find_col(["LastDetected", "LastDetec", "LastSeen"]),
            "ResolvedAt": find_col(["ResolvedAt", "ResolvedDate", "FixedDate", "ClosedDate"]),
            "Resolution": find_col(["Resolution", "ResolutionStatus"]),
            "LocationPath": find_col(["LocationPath", "Location", "Path", "FilePath"]),
            "Projects": find_col(["Projects", "Project", "Application", "App", "ProjectName"]),
            "Link": find_col(["Link", "URL", "WizURL", "Reference", "ReferenceLink", "DetectionLink"]),
            "WizURL": find_col(["WizURL", "WizLink"]),
            "CloudProvider": find_col(["CloudProvider", "Provider", "Cloud"]),
            "CloudPlatform": find_col(["CloudPlatform", "Platform"]),
            "Namespaces": find_col(["Namespaces", "Namespace", "NS"]),
            "Clusters": find_col(["Clusters", "Cluster", "K8sCluster"]),
            "LOB": find_col(["LOB", "LineOfBusiness", "BusinessUnit"]),
            "SubscriptionId": find_col(["SubscriptionId", "SubscriptionID", "SubID"]),
            "SubscriptionName": find_col(["SubscriptionName", "SubName"]),
            "Tags": find_col(["Tags", "Tag", "Labels"]),
        }
        mp = {k: v for k, v in mp.items() if v is not None}

        ni = []
        ri = df.to_dict(orient="records")

        def gv(row, target_key):
            mapped_col = mp.get(target_key)
            if mapped_col and mapped_col in row:
                val = str(row[mapped_col]).strip()
                if val and val.lower() not in ["", "nan", "none", "na", "null"]:
                    return val
            return ""

        for idx, row in enumerate(ri):
            if is_pivot_or_summary_row(row):
                print(f"Skipping pivot/summary row {idx}")
                continue
            rec = {}
            for k, v in row.items():
                rec[k] = str(v).strip() if v is not None else ""

            rec["UploadBatch"] = dsn
            rec["SourceFormat"] = "CONTAINER"  # Mark as Container/Image format

            issue_id = gv(row, "IssueID")
            rec["IssueID"] = issue_id if issue_id else f"VULN-{idx}"

            display_id = gv(row, "DisplayID")
            if display_id and display_id.upper().startswith("CVE"):
                rec["DisplayID"] = display_id
            elif issue_id and issue_id.upper().startswith("CVE"):
                rec["DisplayID"] = issue_id
            elif display_id:
                rec["DisplayID"] = display_id
            else:
                rec["DisplayID"] = rec["IssueID"]

            sev = gv(row, "Severity")
            if sev:
                sev_lower = sev.lower()
                if "critical" in sev_lower:
                    rec["Severity"] = "Critical"
                elif "high" in sev_lower:
                    rec["Severity"] = "High"
                elif "medium" in sev_lower or "moderate" in sev_lower:
                    rec["Severity"] = "Medium"
                elif "low" in sev_lower:
                    rec["Severity"] = "Low"
                elif "info" in sev_lower:
                    rec["Severity"] = "Info"
                else:
                    rec["Severity"] = sev
            else:
                rec["Severity"] = "Medium"

            status = gv(row, "Status")
            rec["Status"] = status if status else "Open"
            category = gv(row, "Category")
            rec["Category"] = category if category else "Uncategorized"
            rec["Department"] = gv(row, "Department")
            rec["AssignedTo"] = gv(row, "AssignedTo")
            rec["DiscoveredDate"] = gv(row, "DiscoveredDate")

            due = gv(row, "DueDate")
            if due:
                rec["DueDate"] = due
            elif rec["DiscoveredDate"]:
                try:
                    dt = pd.to_datetime(rec["DiscoveredDate"], errors='coerce')
                    if pd.notna(dt):
                        dys = 7 if rec["Severity"] == "Critical" else (30 if rec["Severity"] == "High" else 60)
                        rec["DueDate"] = (dt + pd.Timedelta(days=dys)).strftime("%Y-%m-%d")
                except:
                    rec["DueDate"] = ""
            else:
                rec["DueDate"] = ""

            rec["Name"] = gv(row, "Name")
            rec["DetailedName"] = gv(row, "DetailedName")

            # Get original description or generate AI-like short description
            orig_desc = gv(row, "Description")
            if orig_desc and orig_desc.strip() and orig_desc.lower() not in ["na", "none", ""]:
                rec["Description"] = orig_desc
            else:
                # Generate short 5-7 word description
                rec["Description"] = generate_short_description(
                    rec["Name"],
                    rec["DisplayID"],
                    rec["Severity"],
                    gv(row, "AssetType"),
                    rec["DetailedName"]
                )

            # Generate VulnDescription - short 5-7 word vulnerability description
            rec["VulnDescription"] = generate_short_description(
                rec["Name"],
                rec["DisplayID"],
                rec["Severity"],
                gv(row, "AssetType"),
                rec["DetailedName"]
            )

            rec["AffectedAsset"] = gv(row, "AffectedAsset")
            rec["AssetID"] = gv(row, "AssetID")
            rec["AssetType"] = gv(row, "AssetType")
            rem = gv(row, "RecommendedAction")
            rec["RecommendedAction"] = rem if rem else "No action provided"
            rec["ReferenceLinks"] = gv(row, "Link")
            rec["WizURL"] = gv(row, "WizURL")
            rec["Version"] = gv(row, "Version")
            rec["FixedVersion"] = gv(row, "FixedVersion")
            rec["Score"] = gv(row, "Score")
            rec["CVSSSeverity"] = gv(row, "CVSSSeverity")
            rec["VendorSeverity"] = gv(row, "VendorSeverity")
            rec["NvdSeverity"] = gv(row, "NvdSeverity")
            rec["HasExploit"] = gv(row, "HasExploit")
            rec["HasCisaKev"] = gv(row, "HasCisaKev")
            rec["FindingStatus"] = gv(row, "FindingStatus")
            rec["FirstDetected"] = gv(row, "FirstDetected")
            rec["LastDetected"] = gv(row, "LastDetected")
            rec["ResolvedAt"] = gv(row, "ResolvedAt")
            rec["Resolution"] = gv(row, "Resolution")
            rec["LocationPath"] = gv(row, "LocationPath")
            rec["Projects"] = gv(row, "Projects")
            rec["CloudProvider"] = gv(row, "CloudProvider")
            rec["CloudPlatform"] = gv(row, "CloudPlatform")
            rec["Namespaces"] = gv(row, "Namespaces")
            rec["Clusters"] = gv(row, "Clusters")
            rec["LOB"] = gv(row, "LOB")
            rec["SubscriptionId"] = gv(row, "SubscriptionId")
            rec["SubscriptionName"] = gv(row, "SubscriptionName")
            rec["Tags"] = gv(row, "Tags")

            # Auto-assign POD owner based on subscription name/ID
            if not rec["AssignedTo"] or rec["AssignedTo"] in ["", "NA", "Unassigned"]:
                auto_owner = get_pod_owner(rec.get("SubscriptionName", ""), rec.get("SubscriptionId", ""), rec.get("AffectedAsset", ""), rec.get("Projects", ""))
                if auto_owner:
                    rec["AssignedTo"] = auto_owner

            # Filter: Only include Wynk LOB data (skip if LOB exists and is not Wynk)
            # If LOB is empty, include the data
            lob_value = rec["LOB"].lower().strip() if rec["LOB"] else ""
            if lob_value and lob_value not in ALLOWED_LOB and "wynk" not in lob_value:
                print(f"Skipping row {idx}: LOB={rec['LOB']} (not Wynk)")
                continue  # Skip non-Wynk data

            ni.append(rec)

        attach_file_hash(ni, file_hash)
        insert_records(ni, skip_existing_check=True)

        log_upload_history(dsn, fn, "CONTAINER", len(ni))
        print(f"Processed {len(ni)} rows from sheet '{sheetName}'")
        return {"duplicate": False, "status": "success", "processed_rows": len(ni), "format": "CONTAINER"}
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


# No internet connection required - all analysis is done locally

# Remediation templates for common vulnerability types
REMEDIATION_TEMPLATES = {
    "cve": """OFFLINE ANALYSIS - CVE Remediation Steps:

1. IMMEDIATE ACTIONS:
   - Verify the vulnerability exists in your environment
   - Check if the affected component is in production
   - Assess exposure level (internal/external)

2. PATCH STRATEGY:
   - Check vendor advisory for available patches
   - Test patch in staging environment first
   - Schedule maintenance window for production

3. MITIGATION (if patch not available):
   - Apply network-level controls (firewall rules)
   - Implement WAF rules if web-facing
   - Consider disabling affected functionality temporarily

4. VERIFICATION:
   - Re-scan after remediation
   - Update ticket status
   - Document changes made

Note: This is an offline analysis. For detailed CVE information,
consult your internal security documentation.""",

    "config": """OFFLINE ANALYSIS - Configuration Issue:

1. REVIEW:
   - Compare current config against security baseline
   - Check compliance requirements (CIS, NIST, etc.)

2. REMEDIATION:
   - Update configuration to secure defaults
   - Remove unnecessary services/features
   - Implement least privilege principle

3. HARDENING:
   - Enable logging and monitoring
   - Set up alerts for configuration drift
   - Document approved configuration

4. VALIDATION:
   - Test functionality after changes
   - Verify security controls are effective""",

    "default": """OFFLINE ANALYSIS - General Remediation:

1. ASSESSMENT:
   - Review vulnerability details and impact
   - Identify affected systems and data
   - Determine business criticality

2. REMEDIATION OPTIONS:
   - Apply vendor patches if available
   - Implement compensating controls
   - Update security configurations

3. TESTING:
   - Validate fix in test environment
   - Verify no regression in functionality

4. DOCUMENTATION:
   - Update asset inventory
   - Record remediation actions taken
   - Close tracking ticket with evidence"""
}

@app.post("/api/analyze")
async def av(req: Request):
    """Vulnerability analysis using Local Ollama LLM (100% offline)"""
    try:
        data = await req.json()
        description = data.get('description', '')
        asset = data.get('asset', 'Unknown')
        severity = data.get('severity', 'Medium')
        cve_id = data.get('cve', '')

        # Build prompt for Ollama
        prompt = f"""You are a security analyst. Analyze this vulnerability and provide remediation steps.

Vulnerability: {description}
Asset: {asset}
Severity: {severity}
CVE: {cve_id if cve_id else 'N/A'}

Provide:
1. Risk Assessment (2-3 sentences)
2. Immediate Actions (bullet points)
3. Remediation Steps (numbered list)
4. Verification Steps

Keep response concise and actionable."""

        # Call Local Ollama API (127.0.0.1 only - no internet)
        async with httpx.AsyncClient(timeout=180.0) as client:
            response = await client.post(
                OLLAMA_URL,
                json={
                    "model": OLLAMA_MODEL,
                    "prompt": prompt,
                    "stream": False
                }
            )

            if response.status_code == 200:
                result = response.json()
                ai_response = result.get("response", "No response from Ollama")
                return {"remediation": ai_response}
            else:
                # Fallback to template if Ollama fails
                print(f"Ollama error: {response.status_code}")
                return {"remediation": get_fallback_remediation(description, asset)}

    except httpx.ConnectError:
        print("Ollama not running - using fallback")
        return {"remediation": get_fallback_remediation(data.get('description', ''), data.get('asset', 'Unknown'))}
    except Exception as e:
        import traceback
        err = traceback.format_exc()
        print("======== ANALYZE ERROR ========")
        print(err)
        print("===============================")
        return {"remediation": f"Analysis error: {type(e).__name__}: {str(e)}"}


def get_fallback_remediation(description, asset):
    """Fallback when Ollama is not available"""
    desc_lower = description.lower()
    if 'cve' in desc_lower or 'vulnerability' in desc_lower:
        template = REMEDIATION_TEMPLATES["cve"]
    elif 'config' in desc_lower or 'misconfigur' in desc_lower:
        template = REMEDIATION_TEMPLATES["config"]
    else:
        template = REMEDIATION_TEMPLATES["default"]

    return f"""Asset: {asset}

{template}

---
[Fallback mode - Ollama not available. Start Ollama for AI-powered analysis.]"""


# Security knowledge base for offline agent
SECURITY_KB = {
    "critical": "Critical vulnerabilities require immediate attention. Typical SLA is 7 days. Focus on internet-facing systems first.",
    "high": "High severity issues should be addressed within 30 days. Prioritize based on asset criticality.",
    "patch": "Always test patches in staging before production deployment. Document rollback procedures.",
    "sla": "SLA targets: Critical=7d, High=30d, Medium=60d, Low=90d. Track MTTR for continuous improvement.",
    "remediation": "Follow the principle of least privilege. Document all changes. Verify fixes with rescans.",
    "compliance": "Ensure changes align with organizational security policies and compliance frameworks.",
    "risk": "Calculate risk as: Risk = Likelihood x Impact. Prioritize based on business context.",
}

@app.post("/api/ask-agent")
async def aa(req: Request):
    """General AI assistant powered by Local Ollama LLM (100% offline)"""
    try:
        data = await req.json()
        message = data.get('message', '')
        context = data.get('context', [])
        message_lower = message.lower()

        # Only include security context if user asks about it
        security_keywords = ['vulnerability', 'vulnerabilities', 'security', 'critical', 'high', 'cve',
                            'patch', 'remediation', 'sla', 'overdue', 'risk', 'threat', 'exploit',
                            'dashboard', 'issues', 'findings', 'assets']
        is_security_question = any(kw in message_lower for kw in security_keywords)

        context_summary = ""
        if is_security_question and context:
            critical_count = sum(1 for c in context if c.get('Severity') == 'Critical')
            high_count = sum(1 for c in context if c.get('Severity') == 'High')
            open_count = sum(1 for c in context if c.get('Status', '').lower() not in ['resolved', 'closed', 'fixed'])
            context_summary = f"\n\nDashboard Data:\n- Total: {len(context)}\n- Critical: {critical_count}\n- High: {high_count}\n- Open: {open_count}"

        # Build prompt for Ollama - keep it simple
        if is_security_question and context_summary:
            prompt = f"""{message}

{context_summary}"""
        else:
            prompt = message

        # Call Local Ollama API (127.0.0.1 only - no internet)
        async with httpx.AsyncClient(timeout=180.0) as client:
            response = await client.post(
                OLLAMA_URL,
                json={
                    "model": OLLAMA_MODEL,
                    "prompt": prompt,
                    "stream": False
                }
            )

            if response.status_code == 200:
                result = response.json()
                ai_response = result.get("response", "No response from Ollama")
                return {"reply": ai_response}
            else:
                print(f"Ollama error: {response.status_code}")
                return {"reply": get_fallback_agent_response(message, context)}

    except httpx.ConnectError:
        print("Ollama not running - using fallback")
        return {"reply": get_fallback_agent_response(data.get('message', ''), data.get('context', []))}
    except Exception as e:
        import traceback
        err = traceback.format_exc()
        print("======== AGENT ERROR ========")
        print(err)
        print("=============================")
        return {"reply": f"Error: {type(e).__name__}: {str(e)}"}


def get_fallback_agent_response(message, context):
    """Fallback when Ollama is not available"""
    message_lower = message.lower()
    response_parts = []

    if 'critical' in message_lower or 'urgent' in message_lower:
        response_parts.append(SECURITY_KB["critical"])
    if 'high' in message_lower:
        response_parts.append(SECURITY_KB["high"])
    if 'patch' in message_lower or 'update' in message_lower:
        response_parts.append(SECURITY_KB["patch"])
    if 'sla' in message_lower or 'deadline' in message_lower:
        response_parts.append(SECURITY_KB["sla"])
    if 'fix' in message_lower or 'remediat' in message_lower:
        response_parts.append(SECURITY_KB["remediation"])
    if 'complian' in message_lower or 'audit' in message_lower:
        response_parts.append(SECURITY_KB["compliance"])
    if 'risk' in message_lower or 'priorit' in message_lower:
        response_parts.append(SECURITY_KB["risk"])

    if context:
        critical_count = sum(1 for c in context if c.get('Severity') == 'Critical')
        open_count = sum(1 for c in context if c.get('Status', '').lower() not in ['resolved', 'closed', 'fixed'])
        response_parts.append(f"\nCurrent Status: {len(context)} vulnerabilities, {critical_count} critical, {open_count} open.")

    if not response_parts:
        response_parts.append("I'm your AI assistant. I can help with anything - just ask!\n\n[Ollama not running - Start Ollama for full AI capabilities.]")

    return "\n\n".join(response_parts)


@app.get("/api/ollama-status")
async def ollama_status():
    """Check if Ollama is running"""
    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            response = await client.get("http://127.0.0.1:11434/api/tags")
            if response.status_code == 200:
                data = response.json()
                models = [m.get("name") for m in data.get("models", [])]
                return {
                    "status": "running",
                    "models": models,
                    "active_model": OLLAMA_MODEL
                }
    except:
        pass
    return {"status": "offline", "models": [], "message": "Start Ollama with: ollama serve"}


@app.post("/api/smart-search")
async def smart_search(req: Request):
    """Smart search using Ollama to parse natural language queries"""
    try:
        data = await req.json()
        query = data.get('query', '')

        if not query:
            return {"results": [], "error": "No query provided"}

        db = ldb()
        if not db:
            return {"results": [], "message": "No data in database"}

        # First try local parsing (fast, no Ollama needed)
        query_lower = query.lower()
        filters = {"severity": None, "status": None, "keywords": [], "assignee": None, "format": None}

        # Quick severity detection
        if "critical" in query_lower:
            filters["severity"] = "Critical"
        elif "high" in query_lower:
            filters["severity"] = "High"
        elif "medium" in query_lower:
            filters["severity"] = "Medium"
        elif "low" in query_lower:
            filters["severity"] = "Low"

        # Quick status detection
        if "open" in query_lower or "pending" in query_lower:
            filters["status"] = "open"
        elif "resolved" in query_lower or "closed" in query_lower or "fixed" in query_lower:
            filters["status"] = "resolved"

        # Format detection
        if "vapt" in query_lower:
            filters["format"] = "SAST_DAST"
        elif "cspm" in query_lower or "cloud" in query_lower:
            filters["format"] = "CSPM"
        elif "container" in query_lower:
            filters["format"] = "CONTAINER"

        # Common vulnerability keywords
        vuln_keywords = ["sql", "injection", "xss", "rce", "buffer", "overflow", "auth",
                        "bypass", "dos", "denial", "ssrf", "xxe", "path", "traversal",
                        "log4j", "cve", "exploit", "remote", "code", "execution"]
        for kw in vuln_keywords:
            if kw in query_lower:
                filters["keywords"].append(kw)

        # If no filters detected, use Ollama to parse (slower but smarter)
        use_ollama = not filters["severity"] and not filters["status"] and not filters["keywords"]

        if use_ollama:
            try:
                prompt = f"""Parse this security search query and extract filters.
Query: "{query}"

Return ONLY a JSON object with these fields (use null if not mentioned):
{{"severity": "Critical/High/Medium/Low or null", "status": "open/resolved or null", "keywords": ["list", "of", "keywords"], "assignee": "name or null"}}

Return ONLY the JSON, no explanation."""

                async with httpx.AsyncClient(timeout=60.0) as client:
                    response = await client.post(
                        OLLAMA_URL,
                        json={"model": OLLAMA_MODEL, "prompt": prompt, "stream": False}
                    )
                    if response.status_code == 200:
                        ai_response = response.json().get("response", "")
                        # Try to parse JSON from response
                        import re
                        json_match = re.search(r'\{[^}]+\}', ai_response)
                        if json_match:
                            parsed = json.loads(json_match.group())
                            if parsed.get("severity"):
                                filters["severity"] = parsed["severity"]
                            if parsed.get("status"):
                                filters["status"] = parsed["status"]
                            if parsed.get("keywords"):
                                filters["keywords"].extend(parsed["keywords"])
                            if parsed.get("assignee"):
                                filters["assignee"] = parsed["assignee"]
            except:
                pass  # Fall back to basic search

        # Apply filters to database
        results = []
        for item in db:
            match = True

            # Severity filter
            if filters["severity"] and item.get("Severity", "").lower() != filters["severity"].lower():
                match = False

            # Status filter
            if filters["status"]:
                item_status = item.get("Status", "").lower()
                if filters["status"] == "open" and any(x in item_status for x in ["resolved", "closed", "fixed"]):
                    match = False
                elif filters["status"] == "resolved" and not any(x in item_status for x in ["resolved", "closed", "fixed"]):
                    match = False

            # Format filter
            if filters["format"] and item.get("SourceFormat", "CONTAINER") != filters["format"]:
                match = False

            # Keyword filter
            if filters["keywords"]:
                item_text = f"{item.get('Name', '')} {item.get('Description', '')} {item.get('DisplayID', '')}".lower()
                if not any(kw in item_text for kw in filters["keywords"]):
                    match = False

            # Assignee filter
            if filters["assignee"]:
                if filters["assignee"].lower() not in item.get("AssignedTo", "").lower():
                    match = False

            if match:
                results.append(item)

        return {
            "results": results[:100],  # Limit to 100 results
            "total": len(results),
            "filters_applied": filters,
            "query": query
        }

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return {"results": [], "error": str(e)}


@app.post("/api/trigger-openclaw")
async def tc(req: Request):
    """OpenClaw security analysis tool - 100% LOCAL (uses Ollama on 127.0.0.1)"""
    try:
        data = await req.json()
        query = data.get('query', '')
        vuln_context = data.get('context', {})

        if not query:
            return {"result": "No query provided", "tool": "OpenClaw"}

        # Enhanced OpenClaw prompt with vulnerability context
        prompt = f"""You are OpenClaw, an advanced security analysis and threat intelligence tool.

SECURITY QUERY: {query}

VULNERABILITY CONTEXT:
{json.dumps(vuln_context, indent=2) if vuln_context else 'No specific vulnerability context provided.'}

As OpenClaw, provide:
1. THREAT ANALYSIS: Assess the security implications
2. ATTACK VECTORS: Identify potential exploitation methods
3. RISK SCORE: Rate severity (Critical/High/Medium/Low) with justification
4. REMEDIATION: Specific, actionable fix recommendations
5. DETECTION: How to detect if this has been exploited

Be concise but thorough. Focus on actionable intelligence."""

        # Call Local Ollama API (127.0.0.1 only - no internet)
        async with httpx.AsyncClient(timeout=180.0) as client:
            response = await client.post(
                OLLAMA_URL,
                json={
                    "model": OLLAMA_MODEL,
                    "prompt": prompt,
                    "stream": False
                }
            )

            if response.status_code == 200:
                result = response.json()
                return {"result": result.get("response", "No response"), "tool": "OpenClaw"}

        return {"result": "Ollama not available. Start with: ollama serve", "tool": "OpenClaw"}
    except Exception as e:
        return {"result": f"Error: {str(e)}", "tool": "OpenClaw"}

@app.get("/api/calendar/activity")
async def get_calendar_activity(year: int, month: int):
    if not _is_mongo_available():
        return JSONResponse(status_code=503, content={"error": "MongoDB unavailable"})
    try:
        start_date = datetime(year, month, 1, tzinfo=timezone.utc)
        if month == 12:
            end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end_date = datetime(year, month + 1, 1, tzinfo=timezone.utc)

        vuln_pipeline = [
            {"$match": {"UploadedAt": {"$gte": start_date, "$lt": end_date}}},
            {"$group": {
                "_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$UploadedAt"}},
                "count": {"$sum": 1}
            }}
        ]
        vuln_results = list(issues_collection.aggregate(vuln_pipeline))

        upload_pipeline = [
            {"$match": {"UploadedAt": {"$gte": start_date, "$lt": end_date}}},
            {"$group": {
                "_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$UploadedAt"}},
                "count": {"$sum": 1}
            }}
        ]
        upload_results = list(upload_history_collection.aggregate(upload_pipeline))

        activity = {}
        for r in vuln_results:
            date_str = r["_id"]
            if date_str not in activity:
                activity[date_str] = {"vulnerabilities": 0, "uploads": 0}
            activity[date_str]["vulnerabilities"] = r["count"]

        for r in upload_results:
            date_str = r["_id"]
            if date_str not in activity:
                activity[date_str] = {"vulnerabilities": 0, "uploads": 0}
            activity[date_str]["uploads"] = r["count"]

        return activity
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/calendar/vulnerabilities")
async def get_calendar_vulnerabilities(date: str):
    if not _is_mongo_available():
        return JSONResponse(status_code=503, content={"error": "MongoDB unavailable"})
    try:
        year, month, day = map(int, date.split("-"))
        start_date = datetime(year, month, day, tzinfo=timezone.utc)
        end_date = start_date + timedelta(days=1)
        match_stage = {"$match": {"UploadedAt": {"$gte": start_date, "$lt": end_date}}}

        total = issues_collection.count_documents(match_stage["$match"])

        sev_pipeline = [match_stage, {"$group": {"_id": "$Severity", "count": {"$sum": 1}}}]
        sev_results = list(issues_collection.aggregate(sev_pipeline))
        severity_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Info": 0}
        for r in sev_results:
            sev = r["_id"] or "Medium"
            sev_title = sev.title()
            if sev_title in severity_counts:
                severity_counts[sev_title] += r["count"]
            else:
                severity_counts["Info"] += r["count"]

        fmt_pipeline = [match_stage, {"$group": {"_id": "$SourceFormat", "count": {"$sum": 1}}}]
        fmt_results = list(issues_collection.aggregate(fmt_pipeline))
        format_counts = {"CSPM": 0, "VAPT": 0, "CONTAINER": 0, "SAST_DAST": 0}
        for r in fmt_results:
            fmt = r["_id"] or ""
            fmt_upper = fmt.upper()
            if fmt_upper in ["SAST/DAST", "SAST_DAST"]:
                format_counts["SAST_DAST"] += r["count"]
            elif fmt_upper in format_counts:
                format_counts[fmt_upper] += r["count"]

        return {"date": date, "total": total, "severity": severity_counts, "formats": format_counts}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/calendar/uploads")
async def get_calendar_uploads(date: str):
    if not _is_mongo_available():
        return JSONResponse(status_code=503, content={"error": "MongoDB unavailable"})
    try:
        year, month, day = map(int, date.split("-"))
        start_date = datetime(year, month, day, tzinfo=timezone.utc)
        end_date = start_date + timedelta(days=1)
        
        # Use issues_collection to reliably get datasets that have vulnerabilities
        uploads_pipeline = [
            {"$match": {"UploadedAt": {"$gte": start_date, "$lt": end_date}}},
            {"$group": {
                "_id": {"UploadBatch": "$UploadBatch", "SourceFormat": "$SourceFormat"},
                "RecordCount": {"$sum": 1},
                "UploadedAt": {"$first": "$UploadedAt"}
            }},
            {"$project": {
                "_id": 0,
                "FileName": "$_id.UploadBatch",
                "UploadBatch": "$_id.UploadBatch",
                "SourceFormat": "$_id.SourceFormat",
                "RecordCount": 1,
                "UploadedAt": 1
            }},
            {"$sort": {"UploadedAt": -1}}
        ]
        uploads = list(issues_collection.aggregate(uploads_pipeline))

        for u in uploads:
            if "UploadedAt" in u and isinstance(u["UploadedAt"], datetime):
                u["UploadedAt"] = u["UploadedAt"].isoformat()
        return uploads
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/api/analytics/datasets")
async def get_analytics_datasets(formats: str = None, start_date: str = None, end_date: str = None):
    if not _is_mongo_available():
        return JSONResponse(status_code=503, content={"error": "MongoDB unavailable"})
    try:
        match_query = {}
        if formats:
            format_list = [f.strip() for f in formats.split(',')]
            match_query["SourceFormat"] = {"$in": format_list}
            
        if start_date or end_date:
            date_query = {}
            if start_date:
                ymd = start_date.split("T")[0]
                y, m, d = map(int, ymd.split("-"))
                start_dt = datetime(y, m, d, tzinfo=timezone.utc)
                date_query["$gte"] = start_dt
            if end_date:
                ymd = end_date.split("T")[0]
                y, m, d = map(int, ymd.split("-"))
                end_dt = datetime(y, m, d, tzinfo=timezone.utc) + timedelta(days=1) - timedelta(microseconds=1)
                date_query["$lte"] = end_dt
            match_query["UploadedAt"] = date_query

        uploads = list(upload_history_collection.find(match_query, {"_id": 0}).sort("UploadedAt", -1))
        
        for u in uploads:
            if "UploadedAt" in u and isinstance(u["UploadedAt"], datetime):
                u["UploadedAt"] = u["UploadedAt"].isoformat()
                
        return uploads
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/api/analytics/historical")
async def get_analytics_historical(
    formats: str = None, 
    start_date: str = None, 
    end_date: str = None, 
    upload_batches: str = None,
    mode: str = "Cumulative"
):
    if not _is_mongo_available():
        return JSONResponse(status_code=503, content={"error": "MongoDB unavailable"})
    try:
        match_query = {}
        
        if upload_batches:
            batch_list = [b.strip() for b in upload_batches.split('||')]
            match_query["UploadBatch"] = {"$in": batch_list}
        elif formats:
            format_list = [f.strip() for f in formats.split(',')]
            match_query["SourceFormat"] = {"$in": format_list}
            
        # Get baseline query without date restriction to compute running totals
        base_query = match_query.copy()
            
        pipeline = [
            {"$match": base_query},
            {"$group": {
                "_id": {
                    "year": {"$year": "$UploadedAt"},
                    "month": {"$month": "$UploadedAt"},
                    "day": {"$dayOfMonth": "$UploadedAt"}
                },
                "total": {"$sum": 1},
                "resolved": {
                    "$sum": {
                        "$cond": [{"$in": [{"$toLower": "$Status"}, ["resolved", "closed", "fixed", "mitigated", "accepted", "false positive"]]}, 1, 0]
                    }
                },
                "unresolved": {
                    "$sum": {
                        "$cond": [{"$in": [{"$toLower": "$Status"}, ["resolved", "closed", "fixed", "mitigated", "accepted", "false positive"]]}, 0, 1]
                    }
                }
            }},
            {"$sort": {"_id.year": 1, "_id.month": 1, "_id.day": 1}}
        ]
        
        daily_results = list(issues_collection.aggregate(pipeline))
        
        chart_data = []
        cum_total = 0
        cum_res = 0
        cum_unres = 0
        
        start_date_obj = None
        end_date_obj = None
        if start_date:
            ymd = start_date.split("T")[0]
            y, m, d = map(int, ymd.split("-"))
            start_date_obj = date(y, m, d)
        if end_date:
            ymd = end_date.split("T")[0]
            y, m, d = map(int, ymd.split("-"))
            end_date_obj = date(y, m, d)
            
        date_restricted_query = match_query.copy()
        if start_date or end_date:
            date_query = {}
            if start_date:
                date_query["$gte"] = datetime(start_date_obj.year, start_date_obj.month, start_date_obj.day, tzinfo=timezone.utc)
            if end_date:
                date_query["$lte"] = datetime(end_date_obj.year, end_date_obj.month, end_date_obj.day, tzinfo=timezone.utc) + timedelta(days=1) - timedelta(microseconds=1)
            date_restricted_query["UploadedAt"] = date_query
            
        datasets_count = len(issues_collection.distinct("UploadBatch", date_restricted_query))
        
        period_total = 0
        period_res = 0
        period_unres = 0
        
        # Fill missing days
        if daily_results:
            first_day = datetime(daily_results[0]['_id']['year'], daily_results[0]['_id']['month'], daily_results[0]['_id']['day']).date()
            last_day = datetime(daily_results[-1]['_id']['year'], daily_results[-1]['_id']['month'], daily_results[-1]['_id']['day']).date()
            # If end_date is specified and greater than last_day, extend it
            if end_date_obj and end_date_obj > last_day:
                last_day = end_date_obj
                
            current_day = first_day
            idx = 0
            
            while current_day <= last_day:
                d_year, d_month, d_day = current_day.year, current_day.month, current_day.day
                day_str = f"{d_year:04d}-{d_month:02d}-{d_day:02d}"
                
                daily_total = 0
                daily_res = 0
                daily_unres = 0
                
                if idx < len(daily_results):
                    r = daily_results[idx]
                    if r['_id']['year'] == d_year and r['_id']['month'] == d_month and r['_id']['day'] == d_day:
                        daily_total = r["total"]
                        daily_res = r["resolved"]
                        daily_unres = r["unresolved"]
                        idx += 1
                        
                cum_total += daily_total
                cum_res += daily_res
                cum_unres += daily_unres
                
                in_range = True
                if start_date_obj and current_day < start_date_obj:
                    in_range = False
                if end_date_obj and current_day > end_date_obj:
                    in_range = False
                    
                if in_range:
                    period_total += daily_total
                    period_res += daily_res
                    period_unres += daily_unres
                    
                    if mode.lower() == "cumulative":
                        chart_data.append({
                            "date": day_str,
                            "Total": cum_total,
                            "Resolved": cum_res,
                            "Unresolved": cum_unres,
                            "DailyNew": daily_total,
                            "DailyResolved": daily_res,
                            "DailyUnresolved": daily_unres
                        })
                    else:
                        chart_data.append({
                            "date": day_str,
                            "Total": daily_total,
                            "Resolved": daily_res,
                            "Unresolved": daily_unres,
                            "DailyNew": daily_total,
                            "DailyResolved": daily_res,
                            "DailyUnresolved": daily_unres
                        })
                        
                current_day += timedelta(days=1)
                
        summary = {
            "totalDatasets": datasets_count,
            "totalVulnerabilities": period_total if mode.lower() == "daily" else cum_total,
            "resolved": period_res if mode.lower() == "daily" else cum_res,
            "unresolved": period_unres if mode.lower() == "daily" else cum_unres
        }
        
        return {
            "summary": summary,
            "chartData": chart_data
        }
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/api/analytics/owners")
async def get_analytics_owners(
    formats: str = None, 
    start_date: str = None, 
    end_date: str = None, 
    upload_batches: str = None,
    mode: str = "Cumulative",
    owner: str = None
):
    if not _is_mongo_available():
        return JSONResponse(status_code=503, content={"error": "MongoDB unavailable"})
    try:
        match_query = {}
        
        if upload_batches:
            batch_list = [b.strip() for b in upload_batches.split('||')]
            match_query["UploadBatch"] = {"$in": batch_list}
        elif formats:
            format_list = [f.strip() for f in formats.split(',')]
            match_query["SourceFormat"] = {"$in": format_list}
            
        start_date_obj = None
        end_date_obj = None
        if start_date:
            ymd = start_date.split("T")[0]
            y, m, d = map(int, ymd.split("-"))
            start_date_obj = date(y, m, d)
        if end_date:
            ymd = end_date.split("T")[0]
            y, m, d = map(int, ymd.split("-"))
            end_date_obj = date(y, m, d)
            
        # For daily mode we only want issues IN the date range. For cumulative we include everything up to end_date.
        if mode.lower() == "daily":
            if start_date or end_date:
                date_query = {}
                if start_date:
                    date_query["$gte"] = datetime(start_date_obj.year, start_date_obj.month, start_date_obj.day, tzinfo=timezone.utc)
                if end_date:
                    date_query["$lte"] = datetime(end_date_obj.year, end_date_obj.month, end_date_obj.day, tzinfo=timezone.utc) + timedelta(days=1) - timedelta(microseconds=1)
                match_query["UploadedAt"] = date_query
        else: # cumulative
            if end_date:
                match_query["UploadedAt"] = {"$lte": datetime(end_date_obj.year, end_date_obj.month, end_date_obj.day, tzinfo=timezone.utc) + timedelta(days=1) - timedelta(microseconds=1)}

        if owner:
            match_query["AssignedTo"] = owner

        # If no owner is specified, group by owner to return the bar chart data
        if not owner:
            pipeline = [
                {"$match": match_query},
                {"$group": {
                    "_id": {"$ifNull": ["$AssignedTo", "Unassigned"]},
                    "total": {"$sum": 1},
                    "resolved": {
                        "$sum": {
                            "$cond": [{"$in": [{"$toLower": "$Status"}, ["resolved", "closed", "fixed", "mitigated", "accepted", "false positive"]]}, 1, 0]
                        }
                    },
                    "unresolved": {
                        "$sum": {
                            "$cond": [{"$in": [{"$toLower": "$Status"}, ["resolved", "closed", "fixed", "mitigated", "accepted", "false positive"]]}, 0, 1]
                        }
                    }
                }},
                {"$sort": {"total": -1}}
            ]
            results = list(issues_collection.aggregate(pipeline))
            
            owner_data = []
            for r in results:
                owner_data.append({
                    "Owner": r["_id"],
                    "Resolved": r["resolved"],
                    "Unresolved": r["unresolved"],
                    "Total": r["total"]
                })
            return {"ownerData": owner_data}
            
        # If owner IS specified, we want their timeline (Resolved/Unresolved over time) and their summary cards
        # We need the daily aggregation to build a cumulative or daily timeline for this specific owner
        # Just like historical analytics, we need to run it grouped by day.
        # But we use the base_query (without daily restriction) to build a cumulative timeline.
        
        base_query = match_query.copy()
        if mode.lower() == "daily" and "UploadedAt" in base_query:
            del base_query["UploadedAt"] # Remove the date restriction for timeline generation
            
        pipeline = [
            {"$match": base_query},
            {"$group": {
                "_id": {
                    "year": {"$year": "$UploadedAt"},
                    "month": {"$month": "$UploadedAt"},
                    "day": {"$dayOfMonth": "$UploadedAt"}
                },
                "total": {"$sum": 1},
                "resolved": {
                    "$sum": {
                        "$cond": [{"$in": [{"$toLower": "$Status"}, ["resolved", "closed", "fixed", "mitigated", "accepted", "false positive"]]}, 1, 0]
                    }
                },
                "unresolved": {
                    "$sum": {
                        "$cond": [{"$in": [{"$toLower": "$Status"}, ["resolved", "closed", "fixed", "mitigated", "accepted", "false positive"]]}, 0, 1]
                    }
                },
                "critical": {
                    "$sum": {
                        "$cond": [{"$eq": [{"$toLower": "$Severity"}, "critical"]}, 1, 0]
                    }
                },
                "high": {
                    "$sum": {
                        "$cond": [{"$eq": [{"$toLower": "$Severity"}, "high"]}, 1, 0]
                    }
                }
            }},
            {"$sort": {"_id.year": 1, "_id.month": 1, "_id.day": 1}}
        ]
        
        daily_results = list(issues_collection.aggregate(pipeline))
        
        chart_data = []
        cum_total = cum_res = cum_unres = cum_crit = cum_high = 0
        period_total = period_res = period_unres = period_crit = period_high = 0
        
        if daily_results:
            first_day = datetime(daily_results[0]['_id']['year'], daily_results[0]['_id']['month'], daily_results[0]['_id']['day']).date()
            last_day = datetime(daily_results[-1]['_id']['year'], daily_results[-1]['_id']['month'], daily_results[-1]['_id']['day']).date()
            if end_date_obj and end_date_obj > last_day:
                last_day = end_date_obj
                
            current_day = first_day
            idx = 0
            
            while current_day <= last_day:
                d_year, d_month, d_day = current_day.year, current_day.month, current_day.day
                day_str = f"{d_year:04d}-{d_month:02d}-{d_day:02d}"
                
                d_tot = d_res = d_unres = d_crit = d_high = 0
                if idx < len(daily_results):
                    r = daily_results[idx]
                    if r['_id']['year'] == d_year and r['_id']['month'] == d_month and r['_id']['day'] == d_day:
                        d_tot = r["total"]
                        d_res = r["resolved"]
                        d_unres = r["unresolved"]
                        d_crit = r.get("critical", 0)
                        d_high = r.get("high", 0)
                        idx += 1
                        
                cum_total += d_tot
                cum_res += d_res
                cum_unres += d_unres
                cum_crit += d_crit
                cum_high += d_high
                
                in_range = True
                if start_date_obj and current_day < start_date_obj: in_range = False
                if end_date_obj and current_day > end_date_obj: in_range = False
                
                if in_range:
                    period_total += d_tot
                    period_res += d_res
                    period_unres += d_unres
                    period_crit += d_crit
                    period_high += d_high
                    
                    if mode.lower() == "cumulative":
                        chart_data.append({
                            "date": day_str,
                            "Total": cum_total,
                            "Resolved": cum_res,
                            "Unresolved": cum_unres
                        })
                    else:
                        chart_data.append({
                            "date": day_str,
                            "Total": d_tot,
                            "Resolved": d_res,
                            "Unresolved": d_unres
                        })
                        
                current_day += timedelta(days=1)
                
        summary = {
            "Total": period_total if mode.lower() == "daily" else cum_total,
            "Resolved": period_res if mode.lower() == "daily" else cum_res,
            "Unresolved": period_unres if mode.lower() == "daily" else cum_unres,
            "Critical": period_crit if mode.lower() == "daily" else cum_crit,
            "High": period_high if mode.lower() == "daily" else cum_high,
            "Overdue": 0 # Would require due-date logic, stubbing for now
        }
        
        return {
            "summary": summary,
            "chartData": chart_data
        }
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/api/analytics/compare")
async def compare_datasets(batch1: str, batch2: str):
    if not _is_mongo_available():
        return JSONResponse(status_code=503, content={"error": "MongoDB unavailable"})
    try:
        b1_issues = list(issues_collection.find({"UploadBatch": batch1}))
        b2_issues = list(issues_collection.find({"UploadBatch": batch2}))
        
        def get_key(issue):
            return str(issue.get("IssueID") or issue.get("Vulnerability Name") or issue.get("Title") or issue.get("_id"))
            
        b1_dict = {get_key(i): i for i in b1_issues}
        b2_dict = {get_key(i): i for i in b2_issues}
        
        all_keys = set(b1_dict.keys()).union(set(b2_dict.keys()))
        
        comparison = []
        new_findings = 0
        resolved_findings = 0
        still_open = 0
        no_longer_present = 0
        
        for k in all_keys:
            i1 = b1_dict.get(k)
            i2 = b2_dict.get(k)
            
            s1 = str(i1.get("Status", "Open")).lower() if i1 else "none"
            s2 = str(i2.get("Status", "Open")).lower() if i2 else "none"
            
            is_res1 = any(x in s1 for x in ["resolved", "closed", "fixed", "mitigated", "accepted", "false positive"])
            is_res2 = any(x in s2 for x in ["resolved", "closed", "fixed", "mitigated", "accepted", "false positive"])
            
            status_change = ""
            
            if not i1 and i2:
                status_change = "New Findings"
                new_findings += 1
            elif i1 and not i2:
                status_change = "No Longer Present"
                no_longer_present += 1
            elif not is_res1 and is_res2:
                status_change = "Resolved Findings"
                resolved_findings += 1
            elif not is_res1 and not is_res2:
                status_change = "Still Open"
                still_open += 1
            elif is_res1 and is_res2:
                status_change = "Already Resolved"
            else:
                status_change = "Reopened"
                
            comparison.append({
                "Issue": k,
                "Title": str((i2 or i1).get("Vulnerability Name") or (i2 or i1).get("Title") or k),
                "DatasetA_Status": i1.get("Status", "—") if i1 else "—",
                "DatasetB_Status": i2.get("Status", "—") if i2 else "—",
                "Change": status_change,
                "Severity": (i2 or i1).get("Severity", "Unknown")
            })
            
        return {
            "summary": {
                "Dataset1": batch1,
                "Dataset2": batch2,
                "NewFindings": new_findings,
                "ResolvedFindings": resolved_findings,
                "StillOpen": still_open,
                "NoLongerPresent": no_longer_present,
                "TotalD1": len(b1_issues),
                "TotalD2": len(b2_issues)
            },
            "comparison": comparison
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})
@app.delete("/api/dataset")
async def delete_dataset(batch_id: str):
    if not _is_mongo_available():
        return JSONResponse(status_code=503, content={"error": "MongoDB unavailable"})
    try:
        # Delete from upload history
        upload_history_collection.delete_many({"UploadBatch": batch_id})
        # Delete vulnerabilities belonging to this batch
        result = issues_collection.delete_many({"UploadBatch": batch_id})
        
        return {"success": True, "deleted_count": result.deleted_count}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# Only mount static files if dist folder exists (production mode)
if os.path.exists("dist/assets"):
    app.mount("/assets", StaticFiles(directory="dist/assets"), name="assets")

@app.get("/{fp:path}")
async def sr(fp: str):
    fendralis = os.path.join("dist", fp)
    if os.path.exists(fendralis) and os.path.isfile(fendralis):
        return FileResponse(fendralis)
    # In development, return a simple message or fallback
    if os.path.exists("dist/index.html"):
        return FileResponse("dist/index.html")
    return JSONResponse({"message": "Frontend not built. Run 'npm run dev' for development or 'npm run build' for production."})


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)

import re
from datetime import datetime, timezone

@app.post("/api/ai/remediation")
async def ai_remediation(req: Request):
    """Generate structured AI remediation using Ollama"""
    try:
        data = await req.json()
        issue_id = data.get("IssueID")
        upload_batch = data.get("UploadBatch", "")
        source_format = data.get("SourceFormat", "UNKNOWN")
        vulnerability = data.get("vulnerability", {})
        regenerate = data.get("regenerate", False)

        if not issue_id and not vulnerability.get("Name") and not vulnerability.get("finding_name"):
            return ORJSONResponse(status_code=400, content={"error": "Missing IssueID or equivalent identifier"})
            
        # Fallback cache identifier if IssueID is not explicitly present (it usually is)
        cache_id = issue_id or f"{vulnerability.get('Name') or vulnerability.get('finding_name')}-{vulnerability.get('AffectedAsset') or vulnerability.get('resource_name')}"

        # 1. Check Cache
        if not regenerate and _is_mongo_available():
            cached_result = ai_remediation_cache_collection.find_one({
                "IssueID": cache_id,
                "UploadBatch": upload_batch,
                "SourceFormat": source_format
            })
            if cached_result:
                # Remove ObjectId for JSON serialization
                cached_result.pop("_id", None)
                return ORJSONResponse(content={"result": cached_result, "cached": True})

        # 2. Build Format-Aware Prompt
        context_str = ""
        if source_format == "CSPM":
            keys_to_include = ["finding_name", "account_name", "account_id", "resource_type", "resource_id", "resource_name", "impact", "risk_score", "remediation_type", "Description", "RecommendedAction", "ReferenceLinks"]
            context_dict = {k: vulnerability.get(k) for k in keys_to_include if vulnerability.get(k)}
            context_str = "\n".join(f"{k}: {v}" for k, v in context_dict.items())
            
        elif source_format == "VAPT":
            keys_to_include = ["Vulnerability name", "Vulnerability description", "Solution", "Vulnerability Path", "Vulnerability ID", "Vulnerability family", "CVE Number", "Risk Factor", "Severity", "IP", "Hostname", "Port", "Protocol", "Application Owner"]
            context_dict = {k: vulnerability.get(k) for k in keys_to_include if vulnerability.get(k)}
            context_str = "\n".join(f"{k}: {v}" for k, v in context_dict.items())
            
        elif source_format == "CONTAINER":
            keys_to_include = ["Name", "DetailedName", "AffectedAsset", "Severity", "Version", "FixedVersion", "Description", "SubscriptionName", "ImageID", "Namespaces", "Clusters", "RecommendedAction"]
            context_dict = {k: vulnerability.get(k) for k in keys_to_include if vulnerability.get(k)}
            context_str = "\n".join(f"{k}: {v}" for k, v in context_dict.items())
            
        elif source_format == "SAST_DAST":
            keys_to_include = ["issue_key", "Summary", "ApplicationName", "CriticalityStatus", "ReportedOn", "Ageing", "Assignee", "ApplicationOwner", "Description"]
            context_dict = {k: vulnerability.get(k) for k in keys_to_include if vulnerability.get(k)}
            context_str = "\n".join(f"{k}: {v}" for k, v in context_dict.items())
        else:
            # Fallback for unexpected formats
            context_dict = {k: v for k, v in vulnerability.items() if v and isinstance(v, str) and len(v) < 1000}
            context_str = "\n".join(f"{k}: {v}" for k, v in context_dict.items())

        prompt = f"""You are a senior cybersecurity expert analyzing a {source_format} finding. 
Use ONLY the supplied context. Do not invent missing technical facts. Provide practical and actionable remediation.

CONTEXT:
{context_str}

OUTPUT FORMAT EXACTLY AS FOLLOWS (with exactly these section headers in ALL CAPS, do NOT use markdown headers, just the ALL CAPS words followed by a colon and a newline):

FINDING SUMMARY:
(1-2 sentences explaining what the vulnerability means)

ROOT CAUSE:
(Explain the likely underlying configuration/code/security issue)

SECURITY IMPACT:
(Explain what could happen if the issue remains unresolved)

RECOMMENDED REMEDIATION:
(Provide concrete, actionable steps to fix the issue. Use numbered lists.)

VALIDATION STEPS:
(Explain how the security team can verify that the remediation was applied successfully. Use numbered lists.)

PRIORITY RECOMMENDATION:
(One of: Immediate, High, Medium, Low)"""

        # 3. Call Ollama
        async with httpx.AsyncClient(timeout=60.0) as client:
            try:
                response = await client.post(
                    OLLAMA_URL,
                    json={
                        "model": OLLAMA_MODEL,
                        "prompt": prompt,
                        "stream": False
                    }
                )
            except Exception as e:
                return ORJSONResponse(status_code=503, content={"error": "AI remediation is currently unavailable. Please verify that Ollama is running.", "details": str(e)})

            if response.status_code != 200:
                return ORJSONResponse(status_code=500, content={"error": f"Ollama returned error: {response.status_code}"})

            ai_response = response.json().get("response", "")

        # 4. Parse Response Safely
        sections = {
            "AI_Summary": "",
            "AI_RootCause": "",
            "AI_Impact": "",
            "AI_Remediation": [],
            "AI_Validation": [],
            "AI_Priority": "Unknown"
        }
        
        def extract_section(text, current_header, next_header=None):
            try:
                start = text.index(current_header) + len(current_header)
                if next_header:
                    try:
                        end = text.index(next_header, start)
                        return text[start:end].strip()
                    except ValueError:
                        return text[start:].strip()
                return text[start:].strip()
            except ValueError:
                return ""

        summary = extract_section(ai_response, "FINDING SUMMARY:", "ROOT CAUSE:")
        root_cause = extract_section(ai_response, "ROOT CAUSE:", "SECURITY IMPACT:")
        impact = extract_section(ai_response, "SECURITY IMPACT:", "RECOMMENDED REMEDIATION:")
        remediation_str = extract_section(ai_response, "RECOMMENDED REMEDIATION:", "VALIDATION STEPS:")
        validation_str = extract_section(ai_response, "VALIDATION STEPS:", "PRIORITY RECOMMENDATION:")
        priority_str = extract_section(ai_response, "PRIORITY RECOMMENDATION:")

        # Fallback if strict parsing fails
        if not summary and not root_cause:
            sections["AI_Summary"] = "Ollama returned a non-standard format:\n\n" + ai_response
        else:
            sections["AI_Summary"] = summary
            sections["AI_RootCause"] = root_cause
            sections["AI_Impact"] = impact
            sections["AI_Remediation"] = [r.strip() for r in remediation_str.split("\n") if r.strip()]
            sections["AI_Validation"] = [v.strip() for v in validation_str.split("\n") if v.strip()]
            
            p_match = re.search(r'(Immediate|High|Medium|Low)', priority_str, re.IGNORECASE)
            if p_match:
                sections["AI_Priority"] = p_match.group(1).capitalize()

        # 5. Build Result Object
        result = {
            "IssueID": cache_id,
            "UploadBatch": upload_batch,
            "SourceFormat": source_format,
            **sections,
            "AI_GeneratedAt": datetime.now(timezone.utc).isoformat(),
            "AI_Model": OLLAMA_MODEL
        }

        # 6. Save to MongoDB Cache
        if _is_mongo_available():
            ai_remediation_cache_collection.update_one(
                {"IssueID": cache_id, "UploadBatch": upload_batch, "SourceFormat": source_format},
                {"$set": result},
                upsert=True
            )

        return ORJSONResponse(content={"result": result, "cached": False})

    except Exception as e:
        import traceback
        err = traceback.format_exc()
        print(f"Error in /api/ai/remediation:\n{err}")
        return ORJSONResponse(status_code=500, content={"error": "An error occurred while generating AI remediation.", "details": str(e)})


@app.get("/api/email/generate_excel")
async def generate_email_excel(
    request: Request,
    assigned_to: str = None,
    include_graph: str = None,
    columns: str = None,
    search: str = None,
    search_field: str = None,
    severity: str = None,
    status: str = None,
    source_format: str = None,
    upload_batch: str = None,
    date_from: str = None,
    date_to: str = None,
    is_advanced_search: str = None,
    container_sub_types: str = None
):
    if not _is_mongo_available():
        return Response(content="Database unavailable", status_code=503)

    # Use existing DB query builder
    query = _build_db_query(
        search=search, search_field=search_field, severity=severity, status=status,
        assigned_to=assigned_to, source_format=source_format, upload_batch=upload_batch,
        date_from=date_from, date_to=date_to, is_advanced_search=is_advanced_search,
        container_sub_types=container_sub_types
    )

    try:
        cursor = issues_collection.find(query).sort("UploadedAt", -1)
        records = list(cursor)
        
        total = len(records)
        resolved = 0
        unresolved = 0

        # Create DataFrame
        if not records:
            df = pd.DataFrame(["No data found matching filters for this owner."])
        else:
            # Calculate stats
            resolved_terms = ["resolved", "closed", "fixed", "mitigated", "accepted", "false positive"]
            for rec in records:
                st = str(rec.get("Status", "")).lower()
                if any(t in st for t in resolved_terms):
                    resolved += 1
                else:
                    unresolved += 1
                rec.pop("_id", None)
            
            df = pd.DataFrame(records)
            
            # Filter columns if requested
            if columns:
                requested_cols = [c.strip() for c in columns.split(",") if c.strip()]
                existing_cols = [c for c in requested_cols if c in df.columns]
                if existing_cols:
                    df = df[existing_cols]

        excel_buffer = io.BytesIO()
        
        # Use ExcelWriter with openpyxl engine
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Vulnerabilities')
            
            # Embed graph if requested
            if include_graph == "true" and total > 0:
                workbook = writer.book
                stats_sheet = workbook.create_sheet('Statistics')
                stats_sheet['A1'] = 'Status'
                stats_sheet['B1'] = 'Count'
                stats_sheet['A2'] = 'Resolved'
                stats_sheet['B2'] = resolved
                stats_sheet['A3'] = 'Unresolved'
                stats_sheet['B3'] = unresolved
                
                from openpyxl.chart import BarChart, Reference
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = f"Vulnerability Status - {assigned_to or 'Selected Owner'}"
                chart.y_axis.title = 'Count'
                chart.x_axis.title = 'Status'

                data = Reference(stats_sheet, min_col=2, min_row=1, max_row=3, max_col=2)
                cats = Reference(stats_sheet, min_col=1, min_row=2, max_row=3)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                
                stats_sheet.add_chart(chart, "D2")

        excel_buffer.seek(0)
        
        safe_owner = assigned_to.replace(' ', '_') if assigned_to else 'All'
        headers = {
            'Content-Disposition': f'attachment; filename="Security_Vulnerabilities_{safe_owner}.xlsx"',
            'Access-Control-Expose-Headers': 'Content-Disposition, X-Total-Vulnerabilities, X-Resolved, X-Unresolved',
            'X-Total-Vulnerabilities': str(total),
            'X-Resolved': str(resolved),
            'X-Unresolved': str(unresolved)
        }
        return Response(
            content=excel_buffer.getvalue(),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )
        
    except Exception as e:
        import traceback
        err = traceback.format_exc()
        print(f"Error in generate_email_excel:\n{err}")
        return ORJSONResponse({"error": str(e)}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/email/report  — unified Send Mail endpoint
# Uses EXACTLY the same filter path as /api/export and /api/db.
# Returns a ZIP containing:
#   - report.xlsx  (all filtered records, same as Export View)
#   - graph.png    (Resolved vs Unresolved bar chart, only if include_graph=true)
# ─────────────────────────────────────────────────────────────────────────────

def _generate_resolved_unresolved_graph(
    query: dict,
    graph_mode: str,
    owner_label: str,
    format_label: str,
    date_from: str,
    date_to: str,
) -> bytes:
    """
    Runs a MongoDB aggregation over 'query' to get daily Resolved/Unresolved
    counts (same logic as /api/analytics/historical), then renders a PNG chart
    using matplotlib (with Pillow fallback if matplotlib is unavailable).
    Returns raw PNG bytes.
    """
    resolved_statuses = ["resolved", "closed", "fixed", "mitigated", "accepted", "false positive"]

    pipeline = [
        {"$match": query},
        {
            "$group": {
                "_id": {
                    "year":  {"$year":  "$UploadedAt"},
                    "month": {"$month": "$UploadedAt"},
                    "day":   {"$dayOfMonth": "$UploadedAt"},
                },
                "resolved":   {"$sum": {"$cond": [{"$in": [{"$toLower": "$Status"}, resolved_statuses]}, 1, 0]}},
                "unresolved": {"$sum": {"$cond": [{"$in": [{"$toLower": "$Status"}, resolved_statuses]}, 0, 1]}},
            }
        },
        {"$sort": {"_id.year": 1, "_id.month": 1, "_id.day": 1}},
    ]

    daily_results = list(issues_collection.aggregate(pipeline))

    # Build chart data (daily or cumulative) — same logic as /api/analytics/historical
    chart_rows = []  # [{"date": str, "Resolved": int, "Unresolved": int}]
    cum_res = 0
    cum_unres = 0

    for r in daily_results:
        d = r["_id"]
        day_str = f"{d['year']:04d}-{d['month']:02d}-{d['day']:02d}"
        cum_res   += r["resolved"]
        cum_unres += r["unresolved"]
        if graph_mode.lower() == "cumulative":
            chart_rows.append({"date": day_str, "Resolved": cum_res, "Unresolved": cum_unres})
        else:
            chart_rows.append({"date": day_str, "Resolved": r["resolved"], "Unresolved": r["unresolved"]})

    # Subtitle for the chart
    parts = []
    if format_label and format_label != "All": parts.append(format_label)
    if owner_label:  parts.append(owner_label)
    date_range = ""
    if date_from and date_to: date_range = f"{date_from} – {date_to}"
    elif date_from: date_range = f"from {date_from}"
    elif date_to:   date_range = f"to {date_to}"
    if date_range:  parts.append(date_range)
    subtitle = " | ".join(parts) if parts else "All Data"

    # ── Try matplotlib first (best quality) ──────────────────────────────────
    try:
        import matplotlib
        matplotlib.use("Agg")  # non-interactive backend, safe for servers
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        from matplotlib.dates import DateFormatter
        import numpy as np

        dates = [r["date"] for r in chart_rows]
        resolved_vals   = [r["Resolved"]   for r in chart_rows]
        unresolved_vals = [r["Unresolved"] for r in chart_rows]

        fig, ax = plt.subplots(figsize=(12, 5))
        fig.patch.set_facecolor("#f8fafc")
        ax.set_facecolor("#f8fafc")

        x = range(len(dates))
        w = 0.4
        bars_res   = ax.bar([i - w/2 for i in x], resolved_vals,   width=w, label="Resolved",   color="#22c55e", alpha=0.85)
        bars_unres = ax.bar([i + w/2 for i in x], unresolved_vals, width=w, label="Unresolved", color="#ef4444", alpha=0.85)

        # value labels on bars
        for bar in bars_res:
            if bar.get_height() > 0:
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5,
                        str(int(bar.get_height())), ha="center", va="bottom", fontsize=7, color="#166534")
        for bar in bars_unres:
            if bar.get_height() > 0:
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5,
                        str(int(bar.get_height())), ha="center", va="bottom", fontsize=7, color="#991b1b")

        ax.set_xticks(list(x))
        ax.set_xticklabels(dates, rotation=45, ha="right", fontsize=8)
        ax.set_ylabel("Count", fontsize=10)
        ax.set_title(
            f"Resolved vs Unresolved Vulnerabilities — {graph_mode}",
            fontsize=13, fontweight="bold", pad=12
        )
        ax.text(0.5, 1.02, subtitle, transform=ax.transAxes,
                ha="center", fontsize=9, color="#64748b")
        ax.legend(loc="upper right", fontsize=9)
        ax.grid(axis="y", linestyle="--", alpha=0.4)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        plt.tight_layout()

        buf = io.BytesIO()
        plt.savefig(buf, format="png", dpi=150, bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)
        return buf.read()

    except Exception as mpl_err:
        print(f"[email/report] matplotlib unavailable ({mpl_err}), using PIL fallback")

    # ── PIL fallback — draws a simple bar chart ───────────────────────────────
    try:
        from PIL import Image, ImageDraw, ImageFont

        W, H = 1200, 500
        PAD = 60
        BAR_AREA_W = W - 2 * PAD
        BAR_AREA_H = H - 130

        img = Image.new("RGB", (W, H), "#f8fafc")
        draw = ImageDraw.Draw(img)

        # Title
        draw.text((W // 2, 18), f"Resolved vs Unresolved — {graph_mode}",
                  fill="#0f172a", anchor="mt")
        draw.text((W // 2, 40), subtitle, fill="#64748b", anchor="mt")

        n = len(chart_rows)
        if n == 0:
            draw.text((W // 2, H // 2), "No data available", fill="#94a3b8", anchor="mm")
        else:
            max_val = max(
                max((r["Resolved"]   for r in chart_rows), default=0),
                max((r["Unresolved"] for r in chart_rows), default=0),
                1
            )
            slot_w = BAR_AREA_W / n
            bar_w  = max(4, int(slot_w * 0.35))
            top_y  = PAD + 55
            bot_y  = top_y + BAR_AREA_H

            # Axis
            draw.line([(PAD, top_y), (PAD, bot_y), (W - PAD, bot_y)], fill="#cbd5e1", width=1)

            for idx, row in enumerate(chart_rows):
                cx = int(PAD + (idx + 0.5) * slot_w)

                # Resolved bar
                rh = int((row["Resolved"] / max_val) * BAR_AREA_H)
                draw.rectangle([(cx - bar_w - 1, bot_y - rh), (cx - 1, bot_y)], fill="#22c55e")

                # Unresolved bar
                uh = int((row["Unresolved"] / max_val) * BAR_AREA_H)
                draw.rectangle([(cx + 1, bot_y - uh), (cx + bar_w + 1, bot_y)], fill="#ef4444")

                # Date label (every Nth to avoid overlap)
                step = max(1, n // 20)
                if idx % step == 0:
                    draw.text((cx, bot_y + 5), row["date"][-5:],
                              fill="#64748b", anchor="mt")

            # Legend
            draw.rectangle([(PAD, H - 40), (PAD + 14, H - 26)], fill="#22c55e")
            draw.text((PAD + 18, H - 40), "Resolved", fill="#166534", anchor="lt")
            draw.rectangle([(PAD + 90, H - 40), (PAD + 104, H - 26)], fill="#ef4444")
            draw.text((PAD + 108, H - 40), "Unresolved", fill="#991b1b", anchor="lt")

        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        return buf.read()

    except Exception as pil_err:
        print(f"[email/report] PIL fallback also failed: {pil_err}")
        return b""


@app.post("/api/email/report")
async def generate_email_report(
    request: Request,
    source_format: str = None,
    upload_batch: str = None,
    assigned_to: str = None,
    search: str = None,
    search_field: str = None,
    severity: str = None,
    status: str = None,
    date_from: str = None,
    date_to: str = None,
    is_advanced_search: str = None,
    container_sub_types: str = None,
    include_graph: str = "false",
    graph_mode: str = "Daily",
):
    """
    Unified Send Mail report endpoint.

    Accepts exactly the same filter parameters as /api/export and /api/db.
    Calls _build_db_query() ONCE — the same query is used for both:
      1. Excel generation (all filtered records, no pagination)
      2. Graph generation (MongoDB aggregation for Resolved/Unresolved counts)

    Returns a ZIP archive containing:
      - report.xlsx
      - graph.png  (only if include_graph == 'true')
    """
    if not _is_mongo_available():
        return Response(content="Database unavailable", status_code=503)

    # ── Build the shared filter query ─────────────────────────────────────────
    query = _build_db_query(
        search=search,
        search_field=search_field,
        severity=severity,
        status=status,
        assigned_to=assigned_to,
        source_format=source_format,
        upload_batch=upload_batch,
        date_from=date_from,
        date_to=date_to,
        is_advanced_search=is_advanced_search,
        container_sub_types=container_sub_types,
    )

    try:
        # ── 1. Fetch ALL matching records (no pagination — same as Export View) ─
        cursor = issues_collection.find(query).sort("UploadedAt", -1)
        records = list(cursor)

        total      = len(records)
        resolved   = 0
        unresolved = 0
        resolved_terms = ["resolved", "closed", "fixed", "mitigated", "accepted", "false positive"]

        # ── 2. Build Excel ────────────────────────────────────────────────────
        if not records:
            df = pd.DataFrame([{"message": "No vulnerabilities match the selected filters."}])
        else:
            for rec in records:
                st = str(rec.get("Status", "")).lower()
                if any(t in st for t in resolved_terms):
                    resolved += 1
                else:
                    unresolved += 1
                rec.pop("_id", None)
            df = pd.DataFrame(records)

        xlsx_buf = io.BytesIO()
        with pd.ExcelWriter(xlsx_buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Vulnerabilities")

            # Embed a simple summary + openpyxl bar chart in a Statistics sheet
            if total > 0:
                wb = writer.book
                ws = wb.create_sheet("Statistics")
                ws["A1"] = "Scope"
                ws["B1"] = "Value"
                scope_rows = [
                    ("Format",     source_format or "All"),
                    ("Owner",      assigned_to   or "All Owners"),
                    ("Date From",  date_from     or "—"),
                    ("Date To",    date_to       or "—"),
                    ("Graph Mode", graph_mode),
                    ("Total",      total),
                    ("Resolved",   resolved),
                    ("Unresolved", unresolved),
                ]
                for i, (k, v) in enumerate(scope_rows, start=2):
                    ws.cell(row=i, column=1, value=k)
                    ws.cell(row=i, column=2, value=v)

                # openpyxl embedded bar chart (Resolved vs Unresolved)
                from openpyxl.chart import BarChart, Reference
                ws["D1"] = "Status"
                ws["E1"] = "Count"
                ws["D2"] = "Resolved"
                ws["E2"] = resolved
                ws["D3"] = "Unresolved"
                ws["E3"] = unresolved

                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = f"Vulnerability Status — {assigned_to or 'All Owners'}"
                chart.y_axis.title = "Count"
                chart.x_axis.title = "Status"
                data_ref = Reference(ws, min_col=5, min_row=1, max_row=3)
                cats_ref = Reference(ws, min_col=4, min_row=2, max_row=3)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                ws.add_chart(chart, "G2")

        xlsx_buf.seek(0)
        xlsx_bytes = xlsx_buf.read()

        # ── 3. Generate graph PNG (optional) ──────────────────────────────────
        png_bytes = b""
        if include_graph == "true" and total > 0:
            png_bytes = _generate_resolved_unresolved_graph(
                query=query,
                graph_mode=graph_mode,
                owner_label=assigned_to or "",
                format_label=source_format or "All",
                date_from=date_from or "",
                date_to=date_to or "",
            )

        # ── 4. Pack into ZIP ──────────────────────────────────────────────────
        safe_owner = (assigned_to or "All").replace(" ", "_")
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"Security_Report_{safe_owner}.xlsx", xlsx_bytes)
            if png_bytes:
                zf.writestr(f"Resolved_Unresolved_{safe_owner}_{graph_mode}.png", png_bytes)

        zip_buf.seek(0)

        # ── 5. Return ZIP with stats headers ──────────────────────────────────
        from fastapi.responses import Response as FastResponse
        headers = {
            "Content-Disposition": f'attachment; filename="Security_Report_{safe_owner}.zip"',
            "Access-Control-Expose-Headers": "Content-Disposition, X-Total, X-Resolved, X-Unresolved",
            "X-Total":      str(total),
            "X-Resolved":   str(resolved),
            "X-Unresolved": str(unresolved),
        }
        return Response(
            content=zip_buf.read(),
            media_type="application/zip",
            headers=headers,
        )

    except Exception as e:
        import traceback
        print(f"[api/email/report] Error:\n{traceback.format_exc()}")
        return ORJSONResponse({"error": str(e)}, status_code=500)
